import calendar
import re
import shutil
import tempfile
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth.models import AnonymousUser, User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.db import connection
from django.http import HttpResponse
from django.test import RequestFactory, SimpleTestCase, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from . import services
from .error_views import custom_page_not_found, custom_server_error
from .forms import TradingDataExportForm
from .localization import translate
from .models import CapitalMovement, ServerRefreshStatus, SocialLink, Trade, TradingAccount, TradingPreference


class DashboardAccessTests(TestCase):
    def test_dashboard_requires_authentication(self):
        response = self.client.get(reverse('app:dashboard'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('app:login'), response.url)

    def test_transactions_requires_authentication(self):
        response = self.client.get(reverse('app:transactions'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('app:login'), response.url)

    def test_tools_requires_authentication(self):
        response = self.client.get(reverse('app:tools'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('app:login'), response.url)

    def test_auth_pages_mark_first_input_for_autofocus(self):
        login_response = self.client.get(reverse('app:login'))
        register_response = self.client.get(reverse('app:register'))

        self.assertEqual(login_response.status_code, 200)
        self.assertEqual(register_response.status_code, 200)
        self.assertContains(login_response, 'autofocus', html=False)
        self.assertContains(register_response, 'autofocus', html=False)


@override_settings(APP_TRANSLATION_PROVIDER="builtin")
class ErrorPageViewTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def test_custom_404_page_renders_french_content_for_guest(self):
        request = self.factory.get("/page-introuvable/")
        request.user = AnonymousUser()

        response = custom_page_not_found(request, Exception("missing"))

        self.assertEqual(response.status_code, 404)
        self.assertIn("Page introuvable", response.content.decode("utf-8"))
        self.assertIn(reverse("app:login"), response.content.decode("utf-8"))
        self.assertIn(reverse("app:register"), response.content.decode("utf-8"))

    def test_custom_500_page_renders_english_content_from_cookie(self):
        request = self.factory.get("/boom/")
        request.user = type("AuthenticatedUser", (), {"is_authenticated": True})()
        request.COOKIES["django_language"] = "en"

        response = custom_server_error(request)

        content = response.content.decode("utf-8")
        self.assertEqual(response.status_code, 500)
        self.assertIn("Temporary server issue", content)
        self.assertIn("Back to dashboard", content)
        self.assertIn("Reload page", content)
        self.assertIn('lang="en"', content)


class LocalizationProviderTests(SimpleTestCase):
    @override_settings(APP_TRANSLATION_PROVIDER="google_free")
    @patch("app.localization.translate_with_provider")
    def test_translate_uses_external_provider_when_configured(self, mock_translate_with_provider):
        mock_translate_with_provider.return_value = "Account updated: FTMO."

        value = translate(
            "views.success.account_updated",
            language="en",
            account="FTMO",
        )

        self.assertEqual(value, "Account updated: FTMO.")
        mock_translate_with_provider.assert_called_once_with(
            "Compte mis a jour : FTMO.",
            target_language="en",
            source_language="fr",
        )

    @override_settings(APP_TRANSLATION_PROVIDER="google_free")
    @patch("app.localization.translate_with_provider")
    def test_translate_falls_back_to_builtin_when_provider_returns_none(self, mock_translate_with_provider):
        mock_translate_with_provider.return_value = None

        value = translate("dashboard.header.title", language="en")

        self.assertEqual(value, "Performance dashboard")


class TradingDataExportFormTests(SimpleTestCase):
    def test_export_form_builds_week_scope(self):
        form = TradingDataExportForm(
            data={
                'period': TradingDataExportForm.PERIOD_WEEK,
                'week': '2026-W16',
            },
            language='fr',
        )

        self.assertTrue(form.is_valid(), form.errors)
        scope = form.get_scope()
        self.assertEqual(scope['period'], TradingDataExportForm.PERIOD_WEEK)
        self.assertEqual(scope['start_date'], date(2026, 4, 13))
        self.assertEqual(scope['end_date'], date(2026, 4, 19))
        self.assertEqual(scope['filename_token'], '2026-W16')

    def test_export_form_requires_matching_period_value(self):
        form = TradingDataExportForm(
            data={
                'period': TradingDataExportForm.PERIOD_MONTH,
            },
            language='fr',
        )

        self.assertFalse(form.is_valid())
        self.assertIn('month', form.errors)


@override_settings(APP_TRANSLATION_PROVIDER="builtin")
class TradingJournalApiTests(TestCase):
    def setUp(self):
        self.media_root = tempfile.mkdtemp()
        self.settings_ctx = self.settings(MEDIA_ROOT=self.media_root)
        self.settings_ctx.enable()
        self.user = User.objects.create_user(
            username='mikefx',
            email='mike@example.com',
            password='S3curePass123',
        )
        self.client.force_login(self.user)

    def tearDown(self):
        self.settings_ctx.disable()
        shutil.rmtree(self.media_root, ignore_errors=True)
        super().tearDown()

    def get_active_account(self):
        preferences, _ = TradingPreference.objects.get_or_create(user=self.user)
        account, _ = TradingAccount.objects.get_or_create(
            user=self.user,
            name='Compte principal',
            defaults={
                'capital_base': preferences.capital_base,
                'currency': preferences.currency,
            },
        )
        if preferences.active_account_id != account.id:
            preferences.active_account = account
            preferences.capital_base = account.capital_base
            preferences.currency = account.currency
            preferences.save(update_fields=['active_account', 'capital_base', 'currency'])
        return account

    def create_trade(self):
        account = self.get_active_account()
        return Trade.objects.create(
            user=self.user,
            account=account,
            executed_at=timezone.now(),
            symbol='XAUUSD',
            market='Commodities',
            direction=Trade.Direction.LONG,
            result=Trade.Result.TAKE_PROFIT,
            setup='Breakout',
            entry_price=Decimal('1.1000'),
            rr_ratio=Decimal('2.50'),
            exit_price=Decimal('1.1250'),
            quantity=Decimal('100.00'),
            lot_size=Decimal('1.00'),
            gp_value=Decimal('125.00'),
            fees=Decimal('5.00'),
            risk_amount=Decimal('40.00'),
            risk_percent=Decimal('0.40'),
            capital_base=account.capital_base,
            confidence=4,
            notes='Trade test',
        )

    def make_test_image(self, name='trade.gif'):
        return SimpleUploadedFile(
            name,
            (
                b'GIF89a\x01\x00\x01\x00\x80\x00\x00'
                b'\x00\x00\x00\xff\xff\xff!\xf9\x04'
                b'\x01\x00\x00\x00\x00,\x00\x00\x00'
                b'\x00\x01\x00\x01\x00\x00\x02\x02D\x01\x00;'
            ),
            content_type='image/gif',
        )

    def test_dashboard_api_returns_metrics(self):
        self.create_trade()
        response = self.client.get(reverse('app:dashboard-data'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['summary']['has_data'])
        self.assertEqual(payload['overview']['all_time_trade_count'], 1)
        self.assertEqual(len(payload['metrics']), 5)
        self.assertIn('trade_map', payload['calendar'])
        self.assertEqual(payload['preferences']['default_symbol'], 'XAUUSD')
        self.assertEqual(payload['preferences']['currency'], 'USD')
        self.assertEqual(payload['preferences']['current_capital_formatted'], '$10,095')
        self.assertEqual(payload['demo']['action'], 'hidden')
        self.assertEqual(len(payload['recent_trades']), 1)
        self.assertEqual(len(payload['monthly_trades']), 1)

    def test_dashboard_api_exposes_demo_button_state_for_empty_and_seeded_account(self):
        self.get_active_account()

        empty_response = self.client.get(reverse('app:dashboard-data'))

        self.assertEqual(empty_response.status_code, 200)
        self.assertEqual(empty_response.json()['demo']['action'], 'load')
        self.assertEqual(empty_response.json()['demo']['trade_count'], 0)

        services.seed_demo_trades_for_user(self.user.pk)

        seeded_response = self.client.get(reverse('app:dashboard-data'))

        self.assertEqual(seeded_response.status_code, 200)
        self.assertTrue(seeded_response.json()['demo']['loaded'])
        self.assertEqual(seeded_response.json()['demo']['action'], 'unload')
        self.assertGreater(seeded_response.json()['demo']['trade_count'], 0)

    def test_dashboard_view_repairs_invalid_sqlite_decimal_storage(self):
        if connection.vendor != 'sqlite':
            self.skipTest('SQLite-specific regression.')

        trade = self.create_trade()
        with connection.cursor() as cursor:
            cursor.execute(
                'UPDATE app_trade SET rr_ratio = %s, gp_value = %s WHERE id = %s',
                ['', '', trade.id],
            )

        response = self.client.get(reverse('app:dashboard'))

        self.assertEqual(response.status_code, 200)
        trade.refresh_from_db()
        self.assertIsNone(trade.rr_ratio)
        self.assertIsNone(trade.gp_value)

    def test_dashboard_view_repairs_new_invalid_sqlite_decimal_after_prior_request(self):
        if connection.vendor != 'sqlite':
            self.skipTest('SQLite-specific regression.')

        trade = self.create_trade()
        warmup_response = self.client.get(reverse('app:dashboard'))
        self.assertEqual(warmup_response.status_code, 200)

        with connection.cursor() as cursor:
            cursor.execute(
                'UPDATE app_trade SET rr_ratio = %s, gp_value = %s WHERE id = %s',
                ['', '', trade.id],
            )

        response = self.client.get(reverse('app:dashboard'))

        self.assertEqual(response.status_code, 200)
        trade.refresh_from_db()
        self.assertIsNone(trade.rr_ratio)
        self.assertIsNone(trade.gp_value)

    def test_dashboard_view_repairs_non_quantizable_sqlite_decimal_storage(self):
        if connection.vendor != 'sqlite':
            self.skipTest('SQLite-specific regression.')

        trade = self.create_trade()
        with connection.cursor() as cursor:
            cursor.execute(
                'UPDATE app_trade SET rr_ratio = %s, gp_value = %s WHERE id = %s',
                ['1e1000', '1e1000', trade.id],
            )

        response = self.client.get(reverse('app:dashboard'))

        self.assertEqual(response.status_code, 200)
        trade.refresh_from_db()
        self.assertIsNone(trade.rr_ratio)
        self.assertIsNone(trade.gp_value)

    def test_dashboard_view_repairs_invalid_account_and_preference_capital_base_storage(self):
        if connection.vendor != 'sqlite':
            self.skipTest('SQLite-specific regression.')

        account = self.get_active_account()
        preferences = TradingPreference.objects.get(user=self.user)
        with connection.cursor() as cursor:
            cursor.execute(
                'UPDATE app_tradingaccount SET capital_base = %s WHERE id = %s',
                ['1e1000', account.id],
            )
            cursor.execute(
                'UPDATE app_tradingpreference SET capital_base = %s, default_lot_size = %s WHERE id = %s',
                ['1e1000', '1e1000', preferences.id],
            )

        response = self.client.get(reverse('app:dashboard'))

        self.assertEqual(response.status_code, 200)
        account.refresh_from_db()
        preferences.refresh_from_db()
        self.assertEqual(account.capital_base, Decimal('10000.00'))
        self.assertEqual(preferences.capital_base, Decimal('10000.00'))
        self.assertEqual(preferences.default_lot_size, Decimal('1.00'))

    @patch("app.views.ensure_sqlite_decimal_storage_integrity")
    @patch("app.views.render")
    def test_dashboard_view_retries_once_after_invalid_operation(self, mock_render, mock_repair):
        mock_render.side_effect = [
            InvalidOperation(),
            HttpResponse("ok"),
        ]

        response = self.client.get(reverse('app:dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(mock_render.call_count, 2)
        mock_repair.assert_called_once_with()

    def test_repair_sqlite_decimals_management_command_repairs_extreme_values(self):
        if connection.vendor != 'sqlite':
            self.skipTest('SQLite-specific regression.')

        trade = self.create_trade()
        with connection.cursor() as cursor:
            cursor.execute(
                'UPDATE app_trade SET gp_value = %s WHERE id = %s',
                ['1e1000', trade.id],
            )

        call_command('repair_sqlite_decimals')

        trade.refresh_from_db()
        self.assertIsNone(trade.gp_value)

    def test_dashboard_api_limits_recent_trades_to_five_and_exposes_full_month_list(self):
        account = self.get_active_account()
        now = timezone.now()

        for index in range(7):
            Trade.objects.create(
                user=self.user,
                account=account,
                executed_at=now - timedelta(minutes=(6 - index)),
                symbol='XAUUSD',
                market='Commodities',
                direction=Trade.Direction.LONG,
                result=Trade.Result.TAKE_PROFIT,
                setup=f'Trade {index + 1}',
                entry_price=Decimal('1.1000'),
                rr_ratio=Decimal('1.50'),
                exit_price=Decimal('1.1000'),
                quantity=Decimal('1.00'),
                lot_size=Decimal('1.00'),
                gp_value=Decimal('100.00'),
                fees=Decimal('0.00'),
                risk_amount=Decimal('50.00'),
                risk_percent=Decimal('0.50'),
                capital_base=Decimal('10000.00'),
                confidence=4,
            )

        Trade.objects.create(
            user=self.user,
            account=account,
            executed_at=now - timedelta(days=40),
            symbol='NAS100',
            market='Indices',
            direction=Trade.Direction.SHORT,
            result=Trade.Result.STOP_LOSS,
            setup='Old month trade',
            entry_price=Decimal('1.1000'),
            rr_ratio=Decimal('-1.00'),
            exit_price=Decimal('1.1000'),
            quantity=Decimal('1.00'),
            lot_size=Decimal('1.00'),
            gp_value=Decimal('-50.00'),
            fees=Decimal('0.00'),
            risk_amount=Decimal('50.00'),
            risk_percent=Decimal('0.50'),
            capital_base=Decimal('10000.00'),
            confidence=3,
        )

        response = self.client.get(reverse('app:dashboard-data'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(len(payload['recent_trades']), 5)
        self.assertEqual(len(payload['monthly_trades']), 7)
        self.assertEqual(payload['monthly_trades'][0]['setup'], 'Trade 7')
        self.assertEqual(payload['recent_trades'][0]['setup'], 'Trade 7')
        self.assertEqual(payload['recent_trades'][-1]['setup'], 'Trade 3')

    def test_dashboard_api_exposes_only_current_year_when_no_previous_trades_exist(self):
        self.get_active_account()

        response = self.client.get(reverse('app:dashboard-data'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        current_year = str(timezone.localdate().year)
        self.assertEqual(payload['available_years'], [{'value': current_year, 'label': current_year}])
        self.assertEqual(payload['summary']['selected_year'], current_year)

    def test_dashboard_api_uses_saved_default_dashboard_year_when_available(self):
        account = self.get_active_account()
        current_year = timezone.localdate().year
        previous_year = current_year - 1
        Trade.objects.create(
            user=self.user,
            account=account,
            executed_at=timezone.make_aware(datetime(previous_year, 11, 12, 10, 30)),
            symbol='XAUUSD',
            market='Commodities',
            direction=Trade.Direction.LONG,
            result=Trade.Result.TAKE_PROFIT,
            setup='Previous year trade',
            entry_price=Decimal('1.1000'),
            rr_ratio=Decimal('1.50'),
            exit_price=Decimal('1.1000'),
            quantity=Decimal('1.00'),
            lot_size=Decimal('1.00'),
            gp_value=Decimal('100.00'),
            fees=Decimal('0.00'),
            risk_amount=Decimal('50.00'),
            risk_percent=Decimal('0.50'),
            capital_base=Decimal('10000.00'),
            confidence=4,
        )
        preferences = TradingPreference.objects.get(user=self.user)
        preferences.default_dashboard_year = previous_year
        preferences.save(update_fields=['default_dashboard_year', 'updated_at'])

        response = self.client.get(reverse('app:dashboard-data'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['summary']['selected_year'], str(previous_year))
        self.assertEqual(payload['available_years'][0]['value'], str(current_year))
        self.assertEqual(payload['available_years'][1]['value'], str(previous_year))
        self.assertTrue(all(option['value'].startswith(f'{previous_year}-') for option in payload['available_months']))

    def test_dashboard_api_applies_saved_week_start_day_to_calendar_rows(self):
        account = self.get_active_account()
        Trade.objects.create(
            user=self.user,
            account=account,
            executed_at=timezone.make_aware(datetime(2026, 4, 2, 10, 30)),
            symbol='XAUUSD',
            market='Commodities',
            direction=Trade.Direction.LONG,
            result=Trade.Result.TAKE_PROFIT,
            setup='Week start trade',
            entry_price=Decimal('1.1000'),
            rr_ratio=Decimal('1.50'),
            exit_price=Decimal('1.1000'),
            quantity=Decimal('1.00'),
            lot_size=Decimal('1.00'),
            gp_value=Decimal('100.00'),
            fees=Decimal('0.00'),
            risk_amount=Decimal('50.00'),
            risk_percent=Decimal('0.50'),
            capital_base=Decimal('10000.00'),
            confidence=4,
        )
        preferences = TradingPreference.objects.get(user=self.user)

        preferences.default_week_start_day = calendar.SUNDAY
        preferences.save(update_fields=['default_week_start_day', 'updated_at'])

        sunday_response = self.client.get(
            reverse('app:dashboard-data'),
            {'month': '2026-04', 'year': '2026'},
        )

        self.assertEqual(sunday_response.status_code, 200)
        sunday_payload = sunday_response.json()
        self.assertEqual(sunday_payload['calendar']['week_start_day'], calendar.SUNDAY)
        self.assertEqual(sunday_payload['calendar']['weekday_labels'][0], 'Dim')
        self.assertEqual(sunday_payload['calendar']['rows'][0][0]['iso'], '2026-03-29')

        preferences.default_week_start_day = calendar.MONDAY
        preferences.save(update_fields=['default_week_start_day', 'updated_at'])

        monday_response = self.client.get(
            reverse('app:dashboard-data'),
            {'month': '2026-04', 'year': '2026'},
        )

        self.assertEqual(monday_response.status_code, 200)
        monday_payload = monday_response.json()
        self.assertEqual(monday_payload['calendar']['week_start_day'], calendar.MONDAY)
        self.assertEqual(monday_payload['calendar']['weekday_labels'][0], 'Lun')
        self.assertEqual(monday_payload['calendar']['rows'][0][0]['iso'], '2026-03-30')

    @patch("app.services.timezone.localdate")
    def test_dashboard_api_marks_current_day_in_calendar_payload(self, mock_localdate):
        mock_localdate.return_value = date(2026, 4, 7)

        response = self.client.get(
            reverse('app:dashboard-data'),
            {'month': '2026-04', 'year': '2026'},
        )

        self.assertEqual(response.status_code, 200)
        rows = response.json()['calendar']['rows']
        today_entries = [day for row in rows for day in row if day['is_today']]
        self.assertEqual(len(today_entries), 1)
        self.assertEqual(today_entries[0]['iso'], '2026-04-07')

    def test_settings_view_saves_default_dashboard_year_when_previous_year_exists(self):
        account = self.get_active_account()
        previous_year = timezone.localdate().year - 1
        Trade.objects.create(
            user=self.user,
            account=account,
            executed_at=timezone.make_aware(datetime(previous_year, 9, 5, 8, 0)),
            symbol='NAS100',
            market='Indices',
            direction=Trade.Direction.SHORT,
            result=Trade.Result.STOP_LOSS,
            setup='Archived year trade',
            entry_price=Decimal('1.1000'),
            rr_ratio=Decimal('-1.00'),
            exit_price=Decimal('1.1000'),
            quantity=Decimal('1.00'),
            lot_size=Decimal('1.00'),
            gp_value=Decimal('-50.00'),
            fees=Decimal('0.00'),
            risk_amount=Decimal('50.00'),
            risk_percent=Decimal('0.50'),
            capital_base=Decimal('10000.00'),
            confidence=3,
        )

        response = self.client.post(
            reverse('app:settings'),
            data={
                'action': 'preferences',
                'ui_language': 'fr',
                'default_symbol': 'eurusd',
                'default_direction': 'SHORT',
                'default_setup': 'Asia reversal',
                'default_lot_size': '2.50',
                'default_fees': '3.75',
                'default_confidence': '4',
                'default_dashboard_year': str(previous_year),
                'default_week_start_day': str(calendar.SUNDAY),
                'capital_base': '25000.00',
                'currency': 'EUR',
            },
        )

        self.assertEqual(response.status_code, 200)
        preferences = TradingPreference.objects.get(user=self.user)
        self.assertEqual(preferences.default_dashboard_year, previous_year)

    def test_settings_preferences_update_user_language_and_cookie(self):
        response = self.client.post(
            reverse('app:settings'),
            data={
                'action': 'preferences',
                'ui_language': 'en',
                'default_symbol': 'eurusd',
                'default_direction': 'SHORT',
                'default_setup': 'Asia reversal',
                'default_lot_size': '2.50',
                'default_fees': '3.75',
                'default_confidence': '4',
                'default_dashboard_year': str(timezone.localdate().year),
                'default_week_start_day': str(calendar.SUNDAY),
                'capital_base': '25000.00',
                'currency': 'EUR',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<html lang="en"', html=False)
        self.assertEqual(response.cookies['django_language'].value, 'en')
        self.assertEqual(int(response.cookies['django_language']['max-age']), 31536000)

        preferences = TradingPreference.objects.get(user=self.user)
        self.assertEqual(preferences.ui_language, 'en')

    def test_authenticated_user_language_is_restored_without_cookie(self):
        TradingPreference.objects.update_or_create(
            user=self.user,
            defaults={'ui_language': 'es'},
        )
        client = self.client_class()
        client.force_login(self.user)

        response = client.get(reverse('app:dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<html lang="fr"', html=False)
        self.assertEqual(response.cookies['django_language'].value, 'fr')

    def test_language_is_only_configurable_in_settings_with_fr_and_en(self):
        dashboard_response = self.client.get(reverse('app:dashboard'))
        self.assertEqual(dashboard_response.status_code, 200)
        self.assertNotContains(dashboard_response, 'data-language-switcher', html=False)

        settings_response = self.client.get(reverse('app:settings'))
        self.assertEqual(settings_response.status_code, 200)
        self.assertContains(settings_response, 'name="ui_language"', html=False)
        self.assertContains(settings_response, 'value="fr"', html=False)
        self.assertContains(settings_response, 'value="en"', html=False)
        self.assertNotContains(settings_response, 'value="es"', html=False)
        self.assertNotContains(settings_response, 'value="pt"', html=False)
        self.assertNotContains(settings_response, 'value="ar"', html=False)
        self.assertNotContains(settings_response, 'value="zh-hans"', html=False)

    def test_dashboard_view_renders_weekday_headers_from_saved_week_start_day(self):
        self.get_active_account()
        preferences = TradingPreference.objects.get(user=self.user)
        preferences.default_week_start_day = calendar.TUESDAY
        preferences.save(update_fields=['default_week_start_day', 'updated_at'])

        response = self.client.get(reverse('app:dashboard'))

        self.assertEqual(response.status_code, 200)
        content = response.content.decode('utf-8')
        weekdays_block = re.search(r'<div class="calendar-weekdays">(.*?)</div>', content, re.S)
        self.assertIsNotNone(weekdays_block)
        block = weekdays_block.group(1)
        self.assertLess(block.index('<span>Mar</span>'), block.index('<span>Lun</span>'))
        self.assertLess(block.index('<span>Dim</span>'), block.index('<span>Lun</span>'))

    def test_dashboard_view_does_not_mark_trade_symbol_for_initial_focus(self):
        self.get_active_account()

        response = self.client.get(reverse('app:dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'data-trade-initial-focus', html=False)
        self.assertNotContains(response, 'autofocus', html=False)

    def test_login_uses_saved_user_language_instead_of_request_cookie(self):
        TradingPreference.objects.update_or_create(
            user=self.user,
            defaults={'ui_language': 'en'},
        )
        client = self.client_class()
        client.cookies['django_language'] = 'fr'

        login_response = client.post(
            reverse('app:login'),
            data={
                'username': 'mikefx',
                'password': 'S3curePass123',
            },
        )

        self.assertEqual(login_response.status_code, 302)
        self.assertEqual(login_response['Location'], reverse('app:dashboard'))
        self.assertEqual(login_response.cookies['django_language'].value, 'en')

        preferences = TradingPreference.objects.get(user=self.user)
        self.assertEqual(preferences.ui_language, 'en')

        dashboard_response = client.get(reverse('app:dashboard'))
        self.assertEqual(dashboard_response.status_code, 200)
        self.assertContains(dashboard_response, '<html lang="en"', html=False)

    def test_dashboard_trade_payload_exposes_raw_values_for_edit(self):
        Trade.objects.create(
            user=self.user,
            executed_at=timezone.now(),
            symbol='NAS100',
            market='Indices',
            direction=Trade.Direction.SHORT,
            result=Trade.Result.STOP_LOSS,
            setup='Opening drive',
            entry_price=Decimal('3020.5000'),
            rr_ratio=Decimal('-1.50'),
            exit_price=Decimal('3020.5000'),
            quantity=Decimal('2.00'),
            lot_size=Decimal('2.00'),
            gp_value=Decimal('-150.00'),
            fees=Decimal('0.00'),
            risk_amount=Decimal('100.00'),
            risk_percent=Decimal('1.00'),
            confidence=5,
            notes='Test edition front',
        )

        response = self.client.get(reverse('app:dashboard-data'))

        self.assertEqual(response.status_code, 200)
        trade_payload = response.json()['recent_trades'][0]
        self.assertEqual(trade_payload['direction_code'], 'SHORT')
        self.assertEqual(trade_payload['result_code'], 'STOP_LOSS')
        self.assertEqual(trade_payload['result_label'], 'Stoploss')
        self.assertEqual(trade_payload['entry_price_value'], '3020.5000')
        self.assertEqual(trade_payload['ratio_value'], '1.50')
        self.assertEqual(trade_payload['gp_value_value'], '150')
        self.assertEqual(trade_payload['lot_size_value'], '2')

    def test_dashboard_api_applies_break_even_gp_value_to_capital_metrics(self):
        account = self.get_active_account()
        Trade.objects.create(
            user=self.user,
            account=account,
            executed_at=timezone.now(),
            symbol='XAUUSD',
            market='Commodities',
            direction=Trade.Direction.LONG,
            result=Trade.Result.BREAK_EVEN,
            setup='Runner closed late',
            entry_price=Decimal('3020.5000'),
            rr_ratio=Decimal('0.00'),
            exit_price=Decimal('3020.5000'),
            quantity=Decimal('1.00'),
            lot_size=Decimal('1.00'),
            gp_value=Decimal('80.00'),
            fees=Decimal('0.00'),
            risk_amount=Decimal('0.00'),
            risk_percent=Decimal('0.00'),
            capital_base=Decimal('10000.00'),
            confidence=4,
        )

        response = self.client.get(reverse('app:dashboard-data'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['recent_trades'][0]['capital_change_percent_label'], '+0.80%')
        self.assertEqual(payload['preferences']['current_capital_formatted'], '$10,080')

    def test_trade_creation_api_creates_trade(self):
        TradingPreference.objects.update_or_create(
            user=self.user,
            defaults={
                'capital_base': Decimal('25000.00'),
                'default_symbol': 'XAUUSD',
            },
        )
        payload = {
            'executed_at': '2026-03-20T10:30',
            'symbol': 'xauusd',
            'direction': 'LONG',
            'setup': 'VWAP reclaim',
            'entry_price': '3020.5000',
            'rr_ratio': '2.50',
            'result': 'STOP_LOSS',
            'lot_size': '3.00',
            'gp_value': '625.00',
            'confidence': 5,
            'notes': 'Execution propre',
            'screenshots': [
                self.make_test_image('trade-a.gif'),
                self.make_test_image('trade-b.gif'),
            ],
        }

        response = self.client.post(reverse('app:trade-create'), data=payload)

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.json()['trade']['capital_change_percent_label'], '-2.50%')
        self.assertEqual(response.json()['trade']['screenshot_count'], 2)
        self.assertEqual(len(response.json()['trade']['screenshot_urls']), 2)
        self.assertEqual(len(response.json()['trade']['screenshots']), 2)
        self.assertEqual(Trade.objects.count(), 1)
        trade = Trade.objects.get()
        self.assertEqual(trade.symbol, 'XAUUSD')
        self.assertEqual(trade.market, 'Commodities')
        self.assertEqual(trade.result, Trade.Result.STOP_LOSS)
        self.assertEqual(trade.rr_ratio, Decimal('-2.50'))
        self.assertEqual(trade.gp_value, Decimal('-625.00'))
        self.assertEqual(trade.fees, Decimal('0.00'))
        self.assertEqual(trade.capital_base, Decimal('25000.00'))
        self.assertEqual(trade.risk_amount, Decimal('250.00'))
        self.assertEqual(trade.risk_percent, Decimal('1.00'))
        self.assertEqual(trade.screenshots.count(), 2)

    def test_trade_creation_api_allows_positive_break_even_capital_gain(self):
        TradingPreference.objects.update_or_create(
            user=self.user,
            defaults={
                'capital_base': Decimal('10000.00'),
                'default_symbol': 'XAUUSD',
            },
        )
        payload = {
            'executed_at': '2026-03-20T10:30',
            'symbol': 'xauusd',
            'direction': 'LONG',
            'setup': 'Partial runner',
            'entry_price': '3020.5000',
            'rr_ratio': '1.20',
            'result': 'BREAK_EVEN',
            'lot_size': '1.00',
            'gp_value': '150.00',
            'confidence': 4,
            'notes': 'Break even positif',
        }

        response = self.client.post(reverse('app:trade-create'), data=payload)

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.json()['trade']['capital_change_percent_label'], '+1.50%')
        trade = Trade.objects.get(symbol='XAUUSD', setup='Partial runner')
        self.assertEqual(trade.result, Trade.Result.BREAK_EVEN)
        self.assertEqual(trade.rr_ratio, Decimal('0.00'))
        self.assertEqual(trade.gp_value, Decimal('150.00'))
        self.assertEqual(trade.risk_amount, Decimal('0.00'))
        self.assertEqual(trade.risk_percent, Decimal('0.00'))
        self.assertEqual(trade.net_pnl, Decimal('150.00'))

    def test_trade_creation_api_accepts_gain_result(self):
        TradingPreference.objects.update_or_create(
            user=self.user,
            defaults={
                'capital_base': Decimal('10000.00'),
                'default_symbol': 'XAUUSD',
            },
        )
        payload = {
            'executed_at': '2026-03-20T10:30',
            'symbol': 'xauusd',
            'direction': 'LONG',
            'setup': 'Momentum follow-through',
            'entry_price': '3020.5000',
            'rr_ratio': '2.50',
            'result': 'GAIN',
            'lot_size': '1.00',
            'gp_value': '250.00',
            'confidence': 4,
            'notes': 'Gain manuel',
        }

        response = self.client.post(reverse('app:trade-create'), data=payload)

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.json()['trade']['result_code'], 'GAIN')
        self.assertEqual(response.json()['trade']['result_label'], 'Gain')
        self.assertEqual(response.json()['trade']['capital_change_percent_label'], '+2.50%')
        trade = Trade.objects.get(symbol='XAUUSD', setup='Momentum follow-through')
        self.assertEqual(trade.result, Trade.Result.GAIN)
        self.assertEqual(trade.rr_ratio, Decimal('2.50'))
        self.assertEqual(trade.gp_value, Decimal('250.00'))
        self.assertEqual(trade.risk_amount, Decimal('100.00'))

    def test_trade_creation_api_accepts_loss_result(self):
        TradingPreference.objects.update_or_create(
            user=self.user,
            defaults={
                'capital_base': Decimal('10000.00'),
                'default_symbol': 'XAUUSD',
            },
        )
        payload = {
            'executed_at': '2026-03-20T10:30',
            'symbol': 'xauusd',
            'direction': 'LONG',
            'setup': 'Loss manuel',
            'entry_price': '3020.5000',
            'rr_ratio': '2.50',
            'result': 'LOSS',
            'lot_size': '1.00',
            'gp_value': '250.00',
            'confidence': 4,
            'notes': 'Perte manuelle',
        }

        response = self.client.post(reverse('app:trade-create'), data=payload)

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.json()['trade']['result_code'], 'LOSS')
        self.assertEqual(response.json()['trade']['result_label'], 'Perte')
        self.assertEqual(response.json()['trade']['capital_change_percent_label'], '-2.50%')
        trade = Trade.objects.get(symbol='XAUUSD', setup='Loss manuel')
        self.assertEqual(trade.result, Trade.Result.LOSS)
        self.assertEqual(trade.rr_ratio, Decimal('-2.50'))
        self.assertEqual(trade.gp_value, Decimal('-250.00'))
        self.assertEqual(trade.risk_amount, Decimal('100.00'))

    def test_dashboard_api_serializes_trade_gallery(self):
        trade = self.create_trade()
        trade.screenshots.create(image=self.make_test_image('gallery-a.gif'), sort_order=0)
        trade.screenshots.create(image=self.make_test_image('gallery-b.gif'), sort_order=1)

        response = self.client.get(reverse('app:dashboard-data'))

        self.assertEqual(response.status_code, 200)
        trade_payload = response.json()['recent_trades'][0]
        self.assertEqual(trade_payload['screenshot_count'], 2)
        self.assertEqual(len(trade_payload['screenshot_urls']), 2)
        self.assertEqual(len(trade_payload['screenshots']), 2)
        self.assertTrue(trade_payload['screenshot_url'])

    def test_trade_creation_api_uses_current_capital_as_reference(self):
        TradingPreference.objects.update_or_create(
            user=self.user,
            defaults={
                'capital_base': Decimal('10000.00'),
                'default_symbol': 'XAUUSD',
            },
        )
        self.create_trade()
        CapitalMovement.objects.create(
            user=self.user,
            kind=CapitalMovement.Kind.DEPOSIT,
            amount=Decimal('500.00'),
            occurred_at=timezone.now(),
            note='Ajout de capital',
        )
        CapitalMovement.objects.create(
            user=self.user,
            kind=CapitalMovement.Kind.WITHDRAWAL,
            amount=Decimal('200.00'),
            occurred_at=timezone.now(),
            note='Retrait profit',
        )

        response = self.client.post(
            reverse('app:trade-create'),
            data={
                'executed_at': '2026-03-20T10:30',
                'symbol': 'xauusd',
                'direction': 'LONG',
                'setup': 'VWAP reclaim',
                'entry_price': '3020.5000',
                'rr_ratio': '2.50',
                'result': 'STOP_LOSS',
                'lot_size': '3.00',
                'gp_value': '625.00',
                'confidence': 5,
                'notes': 'Execution sur capital actuel',
            },
        )

        self.assertEqual(response.status_code, 201)
        trade = Trade.objects.order_by('-id').first()
        self.assertEqual(trade.capital_base, Decimal('10395.00'))
        self.assertEqual(trade.risk_amount, Decimal('250.00'))
        self.assertEqual(trade.risk_percent, Decimal('2.41'))
        self.assertEqual(response.json()['trade']['capital_change_percent_label'], '-6.01%')

    def test_demo_seed_endpoint_populates_when_empty(self):
        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])
        response = self.client.post(reverse('app:demo-seed'))

        self.assertEqual(response.status_code, 200)
        self.assertGreaterEqual(Trade.objects.count(), 10)

    def test_demo_seed_endpoint_is_restricted_to_super_admin(self):
        response = self.client.post(reverse('app:demo-seed'))

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.json()['ok'], False)

    def test_demo_clear_endpoint_removes_seeded_dataset(self):
        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])
        services.seed_demo_trades_for_user(self.user.pk)

        response = self.client.post(reverse('app:demo-clear'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['ok'], True)
        self.assertEqual(Trade.objects.count(), 0)

    def test_demo_clear_endpoint_is_restricted_to_super_admin(self):
        response = self.client.post(reverse('app:demo-clear'))

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.json()['ok'], False)

    def test_dashboard_shows_super_admin_badge_only_for_super_admin(self):
        self.get_active_account()

        response = self.client.get(reverse('app:dashboard'))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'id="demo-button"', html=False)
        self.assertNotContains(response, 'Super admin')

        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])

        response = self.client.get(reverse('app:dashboard'))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'id="demo-button"', html=False)
        self.assertContains(response, 'Super admin')

    def test_settings_view_shows_demo_controls_only_for_super_admin(self):
        self.get_active_account()

        response = self.client.get(reverse('app:settings'))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'Donnees de demonstration')

        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])

        response = self.client.get(reverse('app:settings'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Donnees de demonstration')
        self.assertContains(response, 'Charger des donnees de demonstration')

    def test_settings_view_allows_super_admin_to_seed_and_clear_demo_data(self):
        self.get_active_account()
        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])

        seed_response = self.client.post(
            reverse('app:settings'),
            data={'action': 'demo_seed'},
        )

        self.assertEqual(seed_response.status_code, 200)
        self.assertContains(seed_response, 'Jeu de donnees de demonstration charge.')
        self.assertGreater(Trade.objects.count(), 0)
        self.assertContains(seed_response, 'Decharger les donnees de demonstration')

        clear_response = self.client.post(
            reverse('app:settings'),
            data={'action': 'demo_clear'},
        )

        self.assertEqual(clear_response.status_code, 200)
        self.assertContains(clear_response, 'Jeu de donnees de demonstration decharge.')
        self.assertEqual(Trade.objects.count(), 0)
        self.assertContains(clear_response, 'Charger des donnees de demonstration')

    def test_dashboard_shows_server_refresh_countdown_only_for_super_admin(self):
        self.get_active_account()
        ServerRefreshStatus.objects.create(
            is_enabled=True,
            last_refreshed_at=timezone.now() - timedelta(days=10),
        )

        response = self.client.get(reverse('app:dashboard'))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'id="server-refresh-countdown"', html=False)

        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])

        response = self.client.get(reverse('app:dashboard'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="server-refresh-countdown"', html=False)
        self.assertContains(response, 'Actualisation mensuelle du serveur')

    def test_dashboard_hides_server_refresh_card_when_disabled(self):
        self.get_active_account()
        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])
        ServerRefreshStatus.objects.create(
            is_enabled=False,
            last_refreshed_at=timezone.now() - timedelta(days=10),
        )

        response = self.client.get(reverse('app:dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'id="server-refresh-countdown"', html=False)
        self.assertNotContains(response, 'Actualisation mensuelle du serveur')

    def test_settings_view_allows_super_admin_to_refresh_server_countdown(self):
        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])
        server_refresh = ServerRefreshStatus.objects.create(
            is_enabled=True,
            last_refreshed_at=timezone.now() - timedelta(days=20),
        )
        previous_refresh_at = server_refresh.last_refreshed_at

        response = self.client.post(
            reverse('app:settings'),
            data={'action': 'server_refresh_mark_updated'},
        )

        self.assertEqual(response.status_code, 200)
        server_refresh.refresh_from_db()
        self.assertTrue(server_refresh.is_enabled)
        self.assertGreater(server_refresh.last_refreshed_at, previous_refresh_at)
        self.assertContains(response, 'Le compte a rebours serveur a ete reinitialise pour un mois.')

    def test_settings_view_allows_super_admin_to_disable_and_enable_server_countdown(self):
        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])
        server_refresh = ServerRefreshStatus.objects.create(
            is_enabled=True,
            last_refreshed_at=timezone.now() - timedelta(days=12),
        )

        disable_response = self.client.post(
            reverse('app:settings'),
            data={'action': 'server_refresh_disable'},
        )

        self.assertEqual(disable_response.status_code, 200)
        server_refresh.refresh_from_db()
        self.assertFalse(server_refresh.is_enabled)
        self.assertContains(disable_response, 'Le suivi mensuel du serveur a ete desactive.')

        previous_refresh_at = server_refresh.last_refreshed_at
        enable_response = self.client.post(
            reverse('app:settings'),
            data={'action': 'server_refresh_enable'},
        )

        self.assertEqual(enable_response.status_code, 200)
        server_refresh.refresh_from_db()
        self.assertTrue(server_refresh.is_enabled)
        self.assertGreater(server_refresh.last_refreshed_at, previous_refresh_at)
        self.assertContains(enable_response, 'Le suivi mensuel du serveur a ete reactive pour un nouveau cycle.')

    def test_settings_view_rejects_server_refresh_actions_for_non_super_admin(self):
        server_refresh = ServerRefreshStatus.objects.create(
            is_enabled=True,
            last_refreshed_at=timezone.now() - timedelta(days=5),
        )

        response = self.client.post(
            reverse('app:settings'),
            data={'action': 'server_refresh_disable'},
        )

        self.assertEqual(response.status_code, 200)
        server_refresh.refresh_from_db()
        self.assertTrue(server_refresh.is_enabled)
        self.assertContains(response, 'Action reservee au super administrateur.')

    def test_transactions_api_returns_monthly_history(self):
        self.create_trade()
        Trade.objects.create(
            user=self.user,
            executed_at=timezone.now() - timedelta(days=35),
            symbol='XAUUSD',
            market='Commodities',
            direction=Trade.Direction.LONG,
            result=Trade.Result.STOP_LOSS,
            setup='Failed breakout',
            entry_price=Decimal('1.1000'),
            rr_ratio=Decimal('-1.00'),
            exit_price=Decimal('1.1000'),
            quantity=Decimal('1.00'),
            lot_size=Decimal('1.00'),
            gp_value=Decimal('-50.00'),
            fees=Decimal('0.00'),
            risk_amount=Decimal('50.00'),
            risk_percent=Decimal('0.50'),
            confidence=3,
            notes='Mois faible',
        )
        CapitalMovement.objects.create(
            user=self.user,
            kind=CapitalMovement.Kind.DEPOSIT,
            amount=Decimal('500.00'),
            occurred_at=timezone.now(),
            note='Ajout de capital',
        )
        CapitalMovement.objects.create(
            user=self.user,
            kind=CapitalMovement.Kind.WITHDRAWAL,
            amount=Decimal('200.00'),
            occurred_at=timezone.now(),
            note='Retrait profit',
        )

        response = self.client.get(reverse('app:transactions-data'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['summary']['current_capital_label'], '$10,345')
        self.assertEqual(payload['summary']['trade_count_month'], 1)
        self.assertEqual(payload['summary']['winners_month'], 1)
        self.assertEqual(payload['summary']['losers_month'], 0)
        self.assertEqual(payload['highlights']['deposits_label'], '$500')
        self.assertEqual(payload['highlights']['withdrawals_label'], '$200')
        self.assertIn('$395', payload['highlights']['best_month_label'])
        self.assertIn('-$50', payload['highlights']['worst_month_label'])
        self.assertEqual(len(payload['monthly_history']), 2)
        self.assertEqual(len(payload['all_movements']), 2)
        self.assertEqual(payload['all_movements'][0]['kind'], 'WITHDRAWAL')
        self.assertEqual(payload['all_movements'][1]['kind'], 'DEPOSIT')
        self.assertEqual(payload['monthly_history'][0]['capital_start_label'], '$9,950')
        self.assertEqual(payload['monthly_history'][1]['capital_start_label'], '$10,000')
        self.assertEqual(payload['monthly_history'][0]['trade_count'], 1)
        self.assertEqual(payload['monthly_history'][0]['winners'], 1)
        self.assertEqual(payload['monthly_history'][0]['losers'], 0)
        self.assertEqual(payload['monthly_history'][0]['progress_label'], '+3.97%')
        self.assertEqual(payload['monthly_history'][0]['progress_tone'], 'profit')
        self.assertEqual(payload['monthly_history'][1]['progress_label'], '-0.50%')
        self.assertEqual(payload['monthly_history'][1]['progress_tone'], 'loss')
        self.assertTrue(payload['monthly_history'][0]['is_best_month'])
        self.assertTrue(payload['monthly_history'][1]['is_worst_month'])

    @patch("app.services.timezone.localdate")
    def test_transactions_api_does_not_mark_current_month_as_worst_month(self, mock_localdate):
        mock_localdate.return_value = datetime(2026, 4, 7).date()
        account = self.get_active_account()

        Trade.objects.create(
            user=self.user,
            account=account,
            executed_at=timezone.make_aware(datetime(2026, 4, 5, 9, 0)),
            symbol='EURUSD',
            market='Forex',
            direction=Trade.Direction.LONG,
            result=Trade.Result.STOP_LOSS,
            setup='Pullback',
            entry_price=Decimal('1.1000'),
            rr_ratio=Decimal('-1.00'),
            exit_price=Decimal('1.0950'),
            quantity=Decimal('1.00'),
            lot_size=Decimal('1.00'),
            gp_value=Decimal('-50.00'),
            fees=Decimal('0.00'),
            risk_amount=Decimal('50.00'),
            risk_percent=Decimal('0.50'),
            capital_base=account.capital_base,
            confidence=3,
            notes='Mois courant en baisse',
        )
        Trade.objects.create(
            user=self.user,
            account=account,
            executed_at=timezone.make_aware(datetime(2026, 3, 10, 9, 0)),
            symbol='XAUUSD',
            market='Commodities',
            direction=Trade.Direction.LONG,
            result=Trade.Result.TAKE_PROFIT,
            setup='Breakout',
            entry_price=Decimal('1.1000'),
            rr_ratio=Decimal('2.00'),
            exit_price=Decimal('1.1200'),
            quantity=Decimal('1.00'),
            lot_size=Decimal('1.00'),
            gp_value=Decimal('100.00'),
            fees=Decimal('0.00'),
            risk_amount=Decimal('50.00'),
            risk_percent=Decimal('0.50'),
            capital_base=account.capital_base,
            confidence=4,
            notes='Mois precedent positif',
        )

        response = self.client.get(reverse('app:transactions-data'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['highlights']['worst_month_label'], '--')
        self.assertEqual(payload['monthly_history'][0]['month_key'], '2026-04')
        self.assertFalse(payload['monthly_history'][0]['is_worst_month'])
        self.assertEqual(payload['monthly_history'][1]['month_key'], '2026-03')
        self.assertFalse(payload['monthly_history'][1]['is_worst_month'])
        self.assertTrue(payload['monthly_history'][1]['is_best_month'])

    def test_capital_movement_create_api_creates_withdrawal(self):
        response = self.client.post(
            reverse('app:capital-movement-create'),
            data={
                'kind': CapitalMovement.Kind.WITHDRAWAL,
                'occurred_at': '2026-03-22T12:45',
                'amount': '150.00',
                'note': 'Retrait hebdo',
            },
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(CapitalMovement.objects.count(), 1)
        movement = CapitalMovement.objects.get()
        self.assertEqual(movement.kind, CapitalMovement.Kind.WITHDRAWAL)
        self.assertEqual(movement.amount, Decimal('150.00'))
        self.assertEqual(response.json()['movement']['amount_label'], '-$150')

    def test_trade_update_api_updates_trade(self):
        trade = self.create_trade()
        TradingPreference.objects.update_or_create(
            user=self.user,
            defaults={
                'capital_base': Decimal('25000.00'),
                'default_symbol': 'XAUUSD',
            },
        )

        response = self.client.post(
            reverse('app:trade-update', args=[trade.pk]),
            data={
                'executed_at': '2026-03-22T09:15',
                'symbol': 'nas100',
                'direction': 'SHORT',
                'setup': 'NY reversal',
                'entry_price': '0.0001',
                'rr_ratio': '1.50',
                'result': 'TAKE_PROFIT',
                'lot_size': '2.50',
                'gp_value': '150.00',
                'confidence': '5',
                'notes': 'Trade modifie',
            },
        )

        self.assertEqual(response.status_code, 200)
        trade.refresh_from_db()
        self.assertEqual(trade.symbol, 'NAS100')
        self.assertEqual(trade.direction, Trade.Direction.SHORT)
        self.assertEqual(trade.setup, 'NY reversal')
        self.assertEqual(trade.result, Trade.Result.TAKE_PROFIT)
        self.assertEqual(trade.rr_ratio, Decimal('1.50'))
        self.assertEqual(trade.gp_value, Decimal('150.00'))
        self.assertEqual(trade.capital_base, Decimal('10095.00'))
        self.assertEqual(trade.risk_amount, Decimal('100.00'))
        self.assertEqual(trade.risk_percent, Decimal('0.99'))
        self.assertEqual(response.json()['trade']['capital_change_percent_label'], '+1.49%')

    def test_trade_update_api_can_remove_existing_screenshots(self):
        trade = self.create_trade()
        screenshot_a = trade.screenshots.create(image=self.make_test_image('edit-a.gif'), sort_order=0)
        screenshot_b = trade.screenshots.create(image=self.make_test_image('edit-b.gif'), sort_order=1)

        response = self.client.post(
            reverse('app:trade-update', args=[trade.pk]),
            data={
                'executed_at': '2026-03-22T09:15',
                'symbol': 'xauusd',
                'direction': 'LONG',
                'setup': 'Breakout',
                'entry_price': '0.0001',
                'rr_ratio': '2.50',
                'result': 'TAKE_PROFIT',
                'lot_size': '1.00',
                'gp_value': '125.00',
                'confidence': '4',
                'notes': 'Trade avec galerie mise a jour',
                'removed_screenshot_ids': [str(screenshot_a.pk)],
            },
        )

        self.assertEqual(response.status_code, 200)
        trade.refresh_from_db()
        remaining_screenshots = list(trade.screenshots.order_by('sort_order', 'pk'))
        self.assertEqual(len(remaining_screenshots), 1)
        self.assertEqual(remaining_screenshots[0].pk, screenshot_b.pk)
        self.assertEqual(remaining_screenshots[0].sort_order, 0)
        self.assertEqual(response.json()['trade']['screenshot_count'], 1)
        self.assertEqual(response.json()['trade']['screenshots'][0]['id'], str(screenshot_b.pk))

    def test_dashboard_payload_allows_editing_legacy_trade_without_ratio_and_gp(self):
        account = self.get_active_account()
        trade = Trade.objects.create(
            user=self.user,
            account=account,
            executed_at=timezone.now(),
            symbol='XAUUSD',
            market='Commodities',
            direction=Trade.Direction.LONG,
            result='',
            setup='Legacy setup',
            entry_price=Decimal('1.0000'),
            exit_price=Decimal('6.0000'),
            quantity=Decimal('25.00'),
            lot_size=None,
            gp_value=None,
            fees=Decimal('0.00'),
            risk_amount=Decimal('50.00'),
            risk_percent=None,
            capital_base=account.capital_base,
            confidence=3,
            notes='Legacy trade',
        )

        dashboard_response = self.client.get(reverse('app:dashboard-data'))
        self.assertEqual(dashboard_response.status_code, 200)
        trade_payload = dashboard_response.json()['monthly_trades'][0]

        self.assertEqual(trade_payload['id'], trade.id)
        self.assertEqual(trade_payload['result_code'], Trade.Result.GAIN)
        self.assertEqual(trade_payload['ratio_value'], '2.50')
        self.assertEqual(trade_payload['gp_value_value'], '125')
        self.assertEqual(trade_payload['lot_size_value'], '25')

        update_response = self.client.post(
            reverse('app:trade-update', args=[trade.pk]),
            data={
                'executed_at': trade_payload['executed_at_input'],
                'symbol': trade_payload['symbol'],
                'direction': trade_payload['direction_code'],
                'setup': 'Legacy setup updated',
                'entry_price': trade_payload['entry_price_value'],
                'rr_ratio': trade_payload['ratio_value'],
                'result': trade_payload['result_code'],
                'lot_size': trade_payload['lot_size_value'],
                'gp_value': trade_payload['gp_value_value'],
                'confidence': str(trade.confidence),
                'notes': 'Legacy updated',
            },
        )

        self.assertEqual(update_response.status_code, 200)
        trade.refresh_from_db()
        self.assertEqual(trade.result, Trade.Result.GAIN)
        self.assertEqual(trade.rr_ratio, Decimal('2.50'))
        self.assertEqual(trade.gp_value, Decimal('125.00'))
        self.assertEqual(trade.lot_size, Decimal('25.00'))
        self.assertEqual(trade.quantity, Decimal('25.00'))
        self.assertEqual(trade.notes, 'Legacy updated')

    def test_settings_view_updates_preferences(self):
        response = self.client.post(
            reverse('app:settings'),
            data={
                'action': 'preferences',
                'ui_language': 'fr',
                'default_symbol': 'eurusd',
                'default_direction': 'SHORT',
                'default_setup': 'Asia reversal',
                'default_lot_size': '2.50',
                'default_fees': '3.75',
                'default_confidence': '4',
                'default_dashboard_year': str(timezone.localdate().year),
                'default_week_start_day': str(calendar.MONDAY),
                'capital_base': '25000.00',
                'currency': 'EUR',
            },
        )

        self.assertEqual(response.status_code, 200)
        preferences = TradingPreference.objects.get(user=self.user)
        self.assertEqual(preferences.default_symbol, 'EURUSD')
        self.assertEqual(preferences.default_direction, Trade.Direction.SHORT)
        self.assertEqual(preferences.default_setup, 'Asia reversal')
        self.assertEqual(preferences.ui_language, 'fr')
        self.assertEqual(preferences.capital_base, Decimal('25000.00'))
        self.assertEqual(preferences.currency, TradingPreference.Currency.EUR)
        self.assertEqual(preferences.default_week_start_day, calendar.MONDAY)
        self.assertContains(response, 'Configuration enregistree.')

    def test_settings_view_displays_current_and_initial_capital_metrics(self):
        self.create_trade()
        CapitalMovement.objects.create(
            user=self.user,
            account=self.get_active_account(),
            kind=CapitalMovement.Kind.DEPOSIT,
            amount=Decimal('500.00'),
            occurred_at=timezone.now(),
            note='Ajout de capital',
        )
        CapitalMovement.objects.create(
            user=self.user,
            account=self.get_active_account(),
            kind=CapitalMovement.Kind.WITHDRAWAL,
            amount=Decimal('200.00'),
            occurred_at=timezone.now(),
            note='Retrait profit',
        )

        response = self.client.get(reverse('app:settings'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '$10,395')
        self.assertContains(response, '$10,000')
        self.assertContains(response, '+3.95%')
        self.assertContains(response, 'Capital initial')

    def test_settings_view_exports_complete_excel_workbook(self):
        account = self.get_active_account()
        archived_account = TradingAccount.objects.create(
            user=self.user,
            name='Compte archive export',
            broker='FTMO',
            account_identifier='ARCH-01',
            capital_base=Decimal('5000.00'),
            currency='EUR',
        )
        archived_account.archived_at = timezone.now()
        archived_account.save(update_fields=['archived_at', 'updated_at'])

        trade = self.create_trade()
        trade.screenshot = self.make_test_image('legacy-export.gif')
        trade.save(update_fields=['screenshot', 'updated_at'])
        trade.screenshots.create(image=self.make_test_image('gallery-export.gif'), sort_order=0)

        CapitalMovement.objects.create(
            user=self.user,
            account=account,
            kind=CapitalMovement.Kind.DEPOSIT,
            amount=Decimal('750.00'),
            occurred_at=timezone.now(),
            note='Depot export',
        )

        response = self.client.post(
            reverse('app:settings'),
            data={
                'action': 'export_data',
                'period': TradingDataExportForm.PERIOD_ALL_TIME,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment; filename="zen-trading-export-all-time.xlsx"', response['Content-Disposition'])

        workbook = load_workbook(filename=BytesIO(response.content))

        self.assertEqual(
            workbook.sheetnames,
            ['Resume', 'Utilisateur', 'Preferences', 'Comptes', 'Trades', 'Mouvements', 'Captures'],
        )

        user_rows = list(workbook['Utilisateur'].iter_rows(values_only=True))
        self.assertEqual(user_rows[1][1], self.user.username)

        account_rows = list(workbook['Comptes'].iter_rows(values_only=True))
        self.assertEqual(len(account_rows), 3)
        self.assertTrue(any(row[1] == 'Compte archive export' and row[7] is True for row in account_rows[1:]))

        trade_rows = list(workbook['Trades'].iter_rows(values_only=True))
        self.assertEqual(len(trade_rows), 2)
        self.assertEqual(trade_rows[1][10], 'Breakout')
        self.assertEqual(trade_rows[1][25], 'Trade test')
        self.assertEqual(trade_rows[1][27], 2)

        movement_rows = list(workbook['Mouvements'].iter_rows(values_only=True))
        self.assertEqual(len(movement_rows), 2)
        self.assertEqual(movement_rows[1][7], 'Depot export')

        screenshot_rows = list(workbook['Captures'].iter_rows(values_only=True))
        self.assertEqual(len(screenshot_rows), 3)
        self.assertEqual({row[5] for row in screenshot_rows[1:]}, {'legacy', 'gallery'})

    def test_settings_view_exports_only_selected_month_for_time_series_data(self):
        account = self.get_active_account()
        april_trade = Trade.objects.create(
            user=self.user,
            account=account,
            executed_at=timezone.make_aware(datetime(2026, 4, 10, 9, 15)),
            symbol='XAUUSD',
            market='Commodities',
            direction=Trade.Direction.LONG,
            result=Trade.Result.TAKE_PROFIT,
            setup='Trade avril',
            entry_price=Decimal('1.1000'),
            rr_ratio=Decimal('2.00'),
            exit_price=Decimal('1.1250'),
            quantity=Decimal('1.00'),
            lot_size=Decimal('1.00'),
            gp_value=Decimal('120.00'),
            fees=Decimal('5.00'),
            risk_amount=Decimal('60.00'),
            risk_percent=Decimal('0.60'),
            capital_base=Decimal('10000.00'),
            confidence=4,
        )
        Trade.objects.create(
            user=self.user,
            account=account,
            executed_at=timezone.make_aware(datetime(2026, 3, 18, 14, 0)),
            symbol='NAS100',
            market='Indices',
            direction=Trade.Direction.SHORT,
            result=Trade.Result.STOP_LOSS,
            setup='Trade mars',
            entry_price=Decimal('1.1000'),
            rr_ratio=Decimal('-1.00'),
            exit_price=Decimal('1.0800'),
            quantity=Decimal('1.00'),
            lot_size=Decimal('1.00'),
            gp_value=Decimal('-60.00'),
            fees=Decimal('3.00'),
            risk_amount=Decimal('60.00'),
            risk_percent=Decimal('0.60'),
            capital_base=Decimal('10000.00'),
            confidence=3,
        )

        CapitalMovement.objects.create(
            user=self.user,
            account=account,
            kind=CapitalMovement.Kind.DEPOSIT,
            amount=Decimal('300.00'),
            occurred_at=timezone.make_aware(datetime(2026, 4, 8, 8, 0)),
            note='Mouvement avril',
        )
        CapitalMovement.objects.create(
            user=self.user,
            account=account,
            kind=CapitalMovement.Kind.WITHDRAWAL,
            amount=Decimal('100.00'),
            occurred_at=timezone.make_aware(datetime(2026, 3, 8, 8, 0)),
            note='Mouvement mars',
        )

        response = self.client.post(
            reverse('app:settings'),
            data={
                'action': 'export_data',
                'period': TradingDataExportForm.PERIOD_MONTH,
                'month': '2026-04',
            },
        )

        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(filename=BytesIO(response.content))
        trade_rows = list(workbook['Trades'].iter_rows(values_only=True))
        movement_rows = list(workbook['Mouvements'].iter_rows(values_only=True))
        account_rows = list(workbook['Comptes'].iter_rows(values_only=True))

        self.assertEqual(len(trade_rows), 2)
        self.assertEqual(trade_rows[1][0], april_trade.pk)
        self.assertEqual(trade_rows[1][10], 'Trade avril')

        self.assertEqual(len(movement_rows), 2)
        self.assertEqual(movement_rows[1][7], 'Mouvement avril')

        self.assertEqual(len(account_rows), 2)
        self.assertEqual(account_rows[1][1], account.name)

    def test_settings_view_does_not_autofocus_any_input(self):
        self.get_active_account()

        response = self.client.get(reverse('app:settings'))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'autofocus', html=False)

    def test_transactions_view_does_not_autofocus_movement_inputs(self):
        self.get_active_account()

        response = self.client.get(reverse('app:transactions'))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'autofocus', html=False)
        self.assertContains(response, 'id="movement-modal-dialog"', html=False)
        self.assertContains(response, 'tabindex="-1"', html=False)

    def test_settings_view_changes_password(self):
        response = self.client.post(
            reverse('app:settings'),
            data={
                'action': 'password',
                'old_password': 'S3curePass123',
                'new_password1': 'NouveauPass456!',
                'new_password2': 'NouveauPass456!',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.user.refresh_from_db()
        self.assertTrue(self.user.check_password('NouveauPass456!'))
        self.assertContains(response, 'Mot de passe mis a jour.')

    def test_settings_view_deletes_account(self):
        response = self.client.post(
            reverse('app:settings'),
            data={
                'action': 'delete_account',
                'password': 'S3curePass123',
                'confirmation': 'on',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(User.objects.filter(pk=self.user.pk).exists())

    def test_settings_view_can_add_and_activate_second_trading_account(self):
        self.get_active_account()

        response = self.client.post(
            reverse('app:settings'),
            data={
                'action': 'create_account',
                'create-account-name': 'FTMO Swing',
                'create-account-broker': 'FTMO',
                'create-account-account_identifier': 'ACC-002',
                'create-account-capital_base': '5000.00',
                'create-account-currency': 'EUR',
                'create-account-set_active': 'on',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(TradingAccount.objects.filter(user=self.user).count(), 2)
        preferences = TradingPreference.objects.get(user=self.user)
        self.assertEqual(preferences.active_account.name, 'FTMO Swing')
        self.assertEqual(preferences.currency, 'EUR')
        self.assertContains(response, 'Compte de trading cree.')

    def test_account_switch_api_changes_active_account(self):
        self.get_active_account()
        second_account = TradingAccount.objects.create(
            user=self.user,
            name='Compte prop',
            broker='FundedNext',
            capital_base=Decimal('5000.00'),
            currency='EUR',
        )

        response = self.client.post(
            reverse('app:account-switch'),
            data={'account_id': second_account.id},
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 200)
        preferences = TradingPreference.objects.get(user=self.user)
        self.assertEqual(preferences.active_account_id, second_account.id)
        self.assertEqual(preferences.currency, 'EUR')
        self.assertEqual(response.json()['account']['name'], 'Compte prop')

    def test_settings_view_can_archive_inactive_account(self):
        primary_account = self.get_active_account()
        second_account = TradingAccount.objects.create(
            user=self.user,
            name='FTMO Swing',
            broker='FTMO',
            capital_base=Decimal('5000.00'),
            currency='EUR',
        )

        response = self.client.post(
            reverse('app:settings'),
            data={
                'action': 'archive_account',
                'account_id': second_account.id,
            },
        )

        self.assertEqual(response.status_code, 200)
        second_account.refresh_from_db()
        preferences = TradingPreference.objects.get(user=self.user)
        self.assertIsNotNone(second_account.archived_at)
        self.assertEqual(preferences.active_account_id, primary_account.id)
        self.assertContains(response, 'Compte archive')

    def test_settings_view_can_update_account_information(self):
        active_account = self.get_active_account()

        response = self.client.post(
            reverse('app:settings'),
            data={
                'action': 'update_account',
                'account_id': active_account.id,
                'edit-account-name': 'Compte perso swing',
                'edit-account-broker': 'IC Markets',
                'edit-account-account_identifier': 'ACC-901',
                'edit-account-capital_base': '15000.00',
                'edit-account-currency': 'GBP',
                'edit-account-set_active': 'on',
            },
        )

        self.assertEqual(response.status_code, 200)
        active_account.refresh_from_db()
        preferences = TradingPreference.objects.get(user=self.user)
        self.assertEqual(active_account.name, 'Compte perso swing')
        self.assertEqual(active_account.broker, 'IC Markets')
        self.assertEqual(active_account.account_identifier, 'ACC-901')
        self.assertEqual(active_account.capital_base, Decimal('15000.00'))
        self.assertEqual(active_account.currency, 'GBP')
        self.assertEqual(preferences.active_account_id, active_account.id)
        self.assertEqual(preferences.currency, 'GBP')
        self.assertContains(response, 'Compte mis a jour')

    def test_settings_view_rejects_archiving_active_account(self):
        primary_account = self.get_active_account()
        TradingAccount.objects.create(
            user=self.user,
            name='Compte secondaire',
            broker='FTMO',
            capital_base=Decimal('5000.00'),
            currency='EUR',
        )

        response = self.client.post(
            reverse('app:settings'),
            data={
                'action': 'archive_account',
                'account_id': primary_account.id,
            },
        )

        self.assertEqual(response.status_code, 200)
        primary_account.refresh_from_db()
        self.assertIsNone(primary_account.archived_at)
        self.assertContains(response, 'Le compte actif ne peut pas etre archive.')

    def test_dashboard_api_uses_active_trading_account_only(self):
        primary_account = self.get_active_account()
        Trade.objects.create(
            user=self.user,
            account=primary_account,
            executed_at=timezone.now(),
            symbol='XAUUSD',
            market='Commodities',
            direction=Trade.Direction.LONG,
            result=Trade.Result.TAKE_PROFIT,
            setup='Primary account trade',
            entry_price=Decimal('1.1000'),
            rr_ratio=Decimal('1.00'),
            exit_price=Decimal('1.1000'),
            quantity=Decimal('1.00'),
            lot_size=Decimal('1.00'),
            gp_value=Decimal('100.00'),
            fees=Decimal('0.00'),
            risk_amount=Decimal('100.00'),
            risk_percent=Decimal('1.00'),
            capital_base=Decimal('10000.00'),
            confidence=4,
        )
        second_account = TradingAccount.objects.create(
            user=self.user,
            name='Compte prop',
            broker='FundedNext',
            capital_base=Decimal('5000.00'),
            currency='EUR',
        )
        Trade.objects.create(
            user=self.user,
            account=second_account,
            executed_at=timezone.now(),
            symbol='NAS100',
            market='Indices',
            direction=Trade.Direction.SHORT,
            result=Trade.Result.TAKE_PROFIT,
            setup='Secondary account trade',
            entry_price=Decimal('1.1000'),
            rr_ratio=Decimal('2.00'),
            exit_price=Decimal('1.1000'),
            quantity=Decimal('1.00'),
            lot_size=Decimal('1.00'),
            gp_value=Decimal('300.00'),
            fees=Decimal('0.00'),
            risk_amount=Decimal('150.00'),
            risk_percent=Decimal('3.00'),
            capital_base=Decimal('5000.00'),
            confidence=5,
        )
        preferences = TradingPreference.objects.get(user=self.user)
        preferences.active_account = second_account
        preferences.capital_base = second_account.capital_base
        preferences.currency = second_account.currency
        preferences.save(update_fields=['active_account', 'capital_base', 'currency'])

        response = self.client.get(reverse('app:dashboard-data'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['overview']['all_time_trade_count'], 1)
        self.assertEqual(payload['overview']['best_setup'], 'Secondary account trade')
        self.assertEqual(payload['preferences']['active_account']['name'], 'Compte prop')
        self.assertIn('5,300', payload['preferences']['current_capital_formatted'])

    def test_dashboard_renders_only_three_active_global_social_links(self):
        self.get_active_account()
        SocialLink.objects.create(
            platform=SocialLink.Platform.WHATSAPP,
            url='https://wa.me/242000000000',
            sort_order=0,
            is_active=True,
        )
        SocialLink.objects.create(
            platform=SocialLink.Platform.X,
            url='https://x.com/akili',
            sort_order=1,
            is_active=True,
        )
        SocialLink.objects.create(
            platform=SocialLink.Platform.TELEGRAM,
            url='https://t.me/akili',
            sort_order=2,
            is_active=True,
        )
        SocialLink.objects.create(
            platform=SocialLink.Platform.YOUTUBE,
            url='https://youtube.com/@akili',
            sort_order=3,
            is_active=True,
        )
        SocialLink.objects.create(
            platform=SocialLink.Platform.LINKEDIN,
            url='https://linkedin.com/company/akili',
            sort_order=4,
            is_active=True,
        )
        SocialLink.objects.create(
            platform=SocialLink.Platform.INSTAGRAM,
            url='https://instagram.com/akili',
            sort_order=0,
            is_active=False,
        )

        response = self.client.get(reverse('app:dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'https://wa.me/242000000000')
        self.assertContains(response, 'https://x.com/akili')
        self.assertContains(response, 'https://t.me/akili')
        self.assertNotContains(response, 'https://youtube.com/@akili')
        self.assertNotContains(response, 'https://linkedin.com/company/akili')
        self.assertNotContains(response, 'https://instagram.com/akili')

    def test_tools_view_renders_core_calculators(self):
        self.get_active_account()

        response = self.client.get(reverse('app:tools'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Outils de trading')
        self.assertContains(response, 'Calculateur de taille de position')
        self.assertContains(response, 'Calculateur de risque de ruine')
