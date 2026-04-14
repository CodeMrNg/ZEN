import calendar
import random
from decimal import Context
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from decimal import Decimal, InvalidOperation
from statistics import pstdev
from threading import Lock
from urllib.parse import urljoin

from django.apps import apps
from django.contrib.auth import get_user_model
from django.core.exceptions import FieldDoesNotExist
from django.core.validators import MinValueValidator
from django.db import connection, models, transaction
from django.db.models.fields import NOT_PROVIDED
from django.utils import timezone
from django.utils.dateparse import parse_date

from .formatting import format_decimal_compact
from .localization import (
    get_rotated_weekday_short_labels,
    normalize_language,
    normalize_week_start_day,
    translate,
)
from .models import (
    CURRENCY_SYMBOLS,
    CapitalMovement,
    ServerRefreshStatus,
    Trade,
    TradeScreenshot,
    TradingAccount,
    TradingPreference,
)

MONTH_LABELS = {
    'fr': {
        1: 'janvier',
        2: 'fevrier',
        3: 'mars',
        4: 'avril',
        5: 'mai',
        6: 'juin',
        7: 'juillet',
        8: 'aout',
        9: 'septembre',
        10: 'octobre',
        11: 'novembre',
        12: 'decembre',
    },
    'en': {
        1: 'january',
        2: 'february',
        3: 'march',
        4: 'april',
        5: 'may',
        6: 'june',
        7: 'july',
        8: 'august',
        9: 'september',
        10: 'october',
        11: 'november',
        12: 'december',
    },
    'es': {
        1: 'enero',
        2: 'febrero',
        3: 'marzo',
        4: 'abril',
        5: 'mayo',
        6: 'junio',
        7: 'julio',
        8: 'agosto',
        9: 'septiembre',
        10: 'octubre',
        11: 'noviembre',
        12: 'diciembre',
    },
    'pt': {
        1: 'janeiro',
        2: 'fevereiro',
        3: 'marco',
        4: 'abril',
        5: 'maio',
        6: 'junho',
        7: 'julho',
        8: 'agosto',
        9: 'setembro',
        10: 'outubro',
        11: 'novembro',
        12: 'dezembro',
    },
}

DEMO_TRADE_NOTE = 'Trade de demonstration genere automatiquement pour illustrer le dashboard.'


_sqlite_decimal_repair_lock = Lock()
_NO_SQLITE_DECIMAL_UPDATE = object()
_SQLITE_DECIMAL_CREATE_CONTEXT = Context(prec=15)
_sqlite_decimal_repair_specs = None


def tr(key, language=None, default=None, **kwargs):
    return translate(key, language=normalize_language(language), default=default, **kwargs)


def format_currency(value, currency='USD'):
    amount = float(value)
    sign = '-' if amount < 0 else ''
    symbol = CURRENCY_SYMBOLS.get(currency, f'{currency} ')
    return f'{sign}{symbol}{format_decimal_compact(abs(value), decimal_places=2, use_grouping=True)}'


def format_signed_value(value):
    amount = float(value)
    sign = '+' if amount > 0 else ''
    return f'{sign}{format_decimal_compact(value, decimal_places=2, use_grouping=True)}'


def format_signed_percent(value):
    amount = float(value)
    sign = '+' if amount > 0 else ''
    return f'{sign}{format_decimal_compact(value, decimal_places=2, use_grouping=True)}%'


def format_local_datetime(value):
    return timezone.localtime(value).strftime('%d/%m/%Y %H:%M')


def format_countdown_compact(total_seconds, language=None):
    code = normalize_language(language)
    labels = {
        'fr': {'day': 'j', 'hour': 'h', 'minute': 'min', 'second': 's'},
        'en': {'day': 'd', 'hour': 'h', 'minute': 'm', 'second': 's'},
        'es': {'day': 'd', 'hour': 'h', 'minute': 'min', 'second': 's'},
        'pt': {'day': 'd', 'hour': 'h', 'minute': 'min', 'second': 's'},
    }.get(code, {'day': 'd', 'hour': 'h', 'minute': 'm', 'second': 's'})

    remaining = max(int(total_seconds), 0)
    days, remainder = divmod(remaining, 86400)
    hours, remainder = divmod(remainder, 3600)
    minutes, seconds = divmod(remainder, 60)

    if days:
        return f"{days} {labels['day']} {hours} {labels['hour']} {minutes} {labels['minute']} {seconds} {labels['second']}"
    if hours:
        return f"{hours} {labels['hour']} {minutes} {labels['minute']} {seconds} {labels['second']}"
    if minutes:
        return f"{minutes} {labels['minute']} {seconds} {labels['second']}"
    return f"{seconds} {labels['second']}"


def get_month_start(value):
    return date(value.year, value.month, 1)


def clamp(value, minimum=0, maximum=100):
    return max(minimum, min(maximum, value))


def _resolve_decimal_repair_fallback(field):
    if field.null:
        return None

    if field.default is not NOT_PROVIDED:
        default_value = field.get_default()
        if default_value not in (None, ''):
            return Decimal(str(default_value)).quantize(
                Decimal(1).scaleb(-field.decimal_places),
                context=field.context,
            )

    for validator in field.validators:
        if isinstance(validator, MinValueValidator) and validator.limit_value is not None:
            return Decimal(str(validator.limit_value)).quantize(
                Decimal(1).scaleb(-field.decimal_places),
                context=field.context,
            )

    return Decimal('0').quantize(
        Decimal(1).scaleb(-field.decimal_places),
        context=field.context,
    )


def _get_sqlite_decimal_repair_specs():
    global _sqlite_decimal_repair_specs

    if _sqlite_decimal_repair_specs is not None:
        return _sqlite_decimal_repair_specs

    repair_specs = []
    for model in apps.get_app_config('app').get_models():
        decimal_fields = []
        for field in model._meta.local_fields:
            if isinstance(field, models.DecimalField):
                decimal_fields.append(
                    {
                        'field': field,
                        'field_name': field.column,
                        'quantize_value': Decimal(1).scaleb(-field.decimal_places),
                        'fallback': _resolve_decimal_repair_fallback(field),
                    }
                )

        if not decimal_fields:
            continue

        try:
            user_column = model._meta.get_field('user').column
        except FieldDoesNotExist:
            user_column = None

        repair_specs.append(
            {
                'model_label': model._meta.label,
                'table_name': model._meta.db_table,
                'pk_column': model._meta.pk.column,
                'user_column': user_column,
                'decimal_fields': decimal_fields,
            }
        )

    _sqlite_decimal_repair_specs = tuple(repair_specs)
    return _sqlite_decimal_repair_specs


def _normalize_sqlite_decimal_replacement(value, field_spec):
    field = field_spec['field']
    fallback = field_spec['fallback']

    if value is None:
        if field.null:
            return _NO_SQLITE_DECIMAL_UPDATE
        return fallback

    raw_value = value.decode('utf-8', errors='ignore') if isinstance(value, bytes) else str(value)
    raw_value = raw_value.strip()
    if not raw_value:
        return fallback

    try:
        normalized = _SQLITE_DECIMAL_CREATE_CONTEXT.create_decimal_from_float(float(raw_value)).quantize(
            field_spec['quantize_value'],
            context=field.context,
        )
    except (InvalidOperation, OverflowError, TypeError, ValueError):
        return fallback

    if not normalized.is_finite():
        return fallback

    normalized_label = f'{normalized:.{field.decimal_places}f}'
    if raw_value == normalized_label:
        return _NO_SQLITE_DECIMAL_UPDATE
    return normalized


def ensure_sqlite_decimal_storage_integrity(user_id=None):
    if connection.vendor != 'sqlite':
        return {'updated_rows': 0, 'updated_fields': 0}

    with _sqlite_decimal_repair_lock:
        if connection.vendor != 'sqlite':
            return {'updated_rows': 0, 'updated_fields': 0}

        existing_tables = set(connection.introspection.table_names())
        updated_rows = 0
        updated_fields = 0
        with transaction.atomic():
            with connection.cursor() as cursor:
                for repair_spec in _get_sqlite_decimal_repair_specs():
                    table_name = repair_spec['table_name']
                    if table_name not in existing_tables:
                        continue

                    pk_column = repair_spec['pk_column']
                    user_column = repair_spec['user_column']

                    for field_spec in repair_spec['decimal_fields']:
                        field_name = field_spec['field_name']
                        query = f'SELECT {pk_column}, {field_name} FROM {table_name}'
                        params = []
                        if user_id is not None and user_column:
                            query += f' WHERE {user_column} = %s'
                            params.append(user_id)

                        cursor.execute(query, params)
                        pending_updates = []
                        for row_id, raw_value in cursor.fetchall():
                            replacement = _normalize_sqlite_decimal_replacement(raw_value, field_spec)
                            if replacement is _NO_SQLITE_DECIMAL_UPDATE:
                                continue
                            pending_updates.append(
                                (None if replacement is None else str(replacement), row_id)
                            )

                        if pending_updates:
                            cursor.executemany(
                                f'UPDATE {table_name} SET {field_name} = %s WHERE {pk_column} = %s',
                                pending_updates,
                            )
                            updated_rows += len(pending_updates)
                            updated_fields += 1

        return {'updated_rows': updated_rows, 'updated_fields': updated_fields}


def format_month_label(year, month, language=None):
    code = normalize_language(language)
    labels = MONTH_LABELS.get(code, MONTH_LABELS['en'])
    return f"{labels[month].capitalize()} {year}"


def format_short_day(day):
    return day.strftime('%d/%m')


def get_available_dashboard_years(all_trades):
    years = {timezone.localdate().year}
    years.update(timezone.localtime(trade.executed_at).year for trade in all_trades)
    return sorted(years, reverse=True)


def resolve_selected_year(raw_year, raw_month, preferences, available_years):
    if raw_month:
        parsed_month = parse_date(f'{raw_month}-01')
        if parsed_month and parsed_month.year in available_years:
            return parsed_month.year

    if raw_year:
        try:
            parsed_year = int(raw_year)
        except (TypeError, ValueError):
            parsed_year = None
        if parsed_year in available_years:
            return parsed_year

    preferred_year = getattr(preferences, 'default_dashboard_year', None) or timezone.localdate().year
    if preferred_year in available_years:
        return preferred_year

    return available_years[0] if available_years else timezone.localdate().year


def resolve_selected_month(raw_month, selected_year, year_trades):
    if raw_month:
        parsed = parse_date(f'{raw_month}-01')
        if parsed and parsed.year == selected_year:
            return parsed.month

    if year_trades:
        latest_trade = max(year_trades, key=lambda trade: trade.executed_at)
        latest_date = timezone.localtime(latest_trade.executed_at).date()
        return latest_date.month

    today = timezone.localdate()
    if selected_year == today.year:
        return today.month
    return 1


def compute_drawdown(cumulative_values):
    peak = 0.0
    max_drawdown = 0.0
    for value in cumulative_values:
        peak = max(peak, value)
        drawdown = value - peak
        max_drawdown = min(max_drawdown, drawdown)
    return max_drawdown


def get_or_create_preferences_for_user(user_id):
    ensure_sqlite_decimal_storage_integrity(user_id=user_id)
    preferences, _ = TradingPreference.objects.get_or_create(user_id=user_id)
    return preferences


def get_or_create_server_refresh_status():
    server_refresh, _ = ServerRefreshStatus.objects.get_or_create(
        pk=1,
        defaults={
            'is_enabled': True,
            'last_refreshed_at': timezone.now(),
        },
    )
    return server_refresh


def enable_server_refresh_tracking():
    server_refresh = get_or_create_server_refresh_status()
    server_refresh.is_enabled = True
    server_refresh.last_refreshed_at = timezone.now()
    server_refresh.save(update_fields=['is_enabled', 'last_refreshed_at', 'updated_at'])
    return server_refresh


def disable_server_refresh_tracking():
    server_refresh = get_or_create_server_refresh_status()
    if server_refresh.is_enabled:
        server_refresh.is_enabled = False
        server_refresh.save(update_fields=['is_enabled', 'updated_at'])
    return server_refresh


def mark_server_refresh_updated():
    server_refresh = get_or_create_server_refresh_status()
    server_refresh.is_enabled = True
    server_refresh.last_refreshed_at = timezone.now()
    server_refresh.save(update_fields=['is_enabled', 'last_refreshed_at', 'updated_at'])
    return server_refresh


def build_server_refresh_snapshot(language=None):
    server_refresh = get_or_create_server_refresh_status()
    due_at = server_refresh.next_refresh_due_at if server_refresh.is_enabled else None
    seconds_remaining = int((due_at - timezone.now()).total_seconds()) if due_at else None
    is_overdue = bool(server_refresh.is_enabled and seconds_remaining is not None and seconds_remaining < 0)

    if not server_refresh.is_enabled:
        status_label = tr('server_refresh.status.disabled', language=language, default='Desactive')
        countdown_label = tr('server_refresh.countdown.disabled', language=language, default='Desactive')
        summary_label = tr(
            'server_refresh.summary.disabled',
            language=language,
            default='Le suivi mensuel de l actualisation serveur est desactive.',
        )
    elif is_overdue:
        overdue_duration = format_countdown_compact(abs(seconds_remaining), language=language)
        status_label = tr('server_refresh.status.overdue', language=language, default='En retard')
        countdown_label = tr(
            'server_refresh.countdown.overdue',
            language=language,
            default='Depasse de {duration}',
            duration=overdue_duration,
        )
        summary_label = tr(
            'server_refresh.summary.overdue',
            language=language,
            default='L actualisation mensuelle du serveur est depassee. Actualisez-la ou desactivez le suivi dans les parametres.',
        )
    else:
        status_label = tr('server_refresh.status.enabled', language=language, default='Suivi actif')
        countdown_label = format_countdown_compact(seconds_remaining, language=language)
        summary_label = tr(
            'server_refresh.summary.enabled',
            language=language,
            default='Prochaine actualisation requise avant le {date}.',
            date=format_local_datetime(due_at),
        )

    return {
        'is_enabled': server_refresh.is_enabled,
        'is_overdue': is_overdue,
        'status_label': status_label,
        'summary_label': summary_label,
        'countdown_label': countdown_label,
        'last_refreshed_at': server_refresh.last_refreshed_at,
        'last_refreshed_at_label': format_local_datetime(server_refresh.last_refreshed_at),
        'due_at': due_at,
        'due_at_iso': timezone.localtime(due_at).isoformat() if due_at else '',
        'due_at_label': format_local_datetime(due_at) if due_at else '--',
    }


def build_account_label(account, language=None):
    if not account:
        return tr('common.main_account', language=language, default='Compte principal')

    parts = [account.name]
    if account.broker:
        parts.append(account.broker)
    return ' | '.join(parts)


def sync_preferences_with_account(preferences, account):
    update_fields = []

    if preferences.active_account_id != account.id:
        preferences.active_account = account
        update_fields.append('active_account')
    if preferences.capital_base != account.capital_base:
        preferences.capital_base = account.capital_base
        update_fields.append('capital_base')
    if preferences.currency != account.currency:
        preferences.currency = account.currency
        update_fields.append('currency')

    if update_fields:
        preferences.save(update_fields=update_fields)

    return preferences


def get_or_create_active_account_for_user(user_id, preferences=None):
    preferences = preferences or get_or_create_preferences_for_user(user_id)

    account = None
    if preferences.active_account_id:
        account = TradingAccount.objects.filter(
            user_id=user_id,
            pk=preferences.active_account_id,
            archived_at__isnull=True,
        ).first()

    if account is None:
        account = TradingAccount.objects.filter(
            user_id=user_id,
            archived_at__isnull=True,
        ).order_by('created_at', 'pk').first()

    if account is None:
        account = TradingAccount.objects.create(
            user_id=user_id,
            name=tr('common.main_account', default='Compte principal'),
            capital_base=preferences.capital_base,
            currency=preferences.currency,
        )

    sync_preferences_with_account(preferences, account)
    return account


def get_trading_accounts_for_user(user_id, include_archived=False):
    queryset = TradingAccount.objects.filter(user_id=user_id)
    if not include_archived:
        queryset = queryset.filter(archived_at__isnull=True)
    return queryset.order_by('created_at', 'pk')


def get_archived_trading_accounts_for_user(user_id):
    return TradingAccount.objects.filter(
        user_id=user_id,
        archived_at__isnull=False,
    ).order_by('-archived_at', '-updated_at', '-pk')


def switch_active_account_for_user(user_id, account_id):
    preferences = get_or_create_preferences_for_user(user_id)
    account = TradingAccount.objects.filter(
        user_id=user_id,
        pk=account_id,
        archived_at__isnull=True,
    ).first()
    if account is None:
        return None

    sync_preferences_with_account(preferences, account)
    return account


def archive_trading_account_for_user(user_id, account_id):
    preferences = get_or_create_preferences_for_user(user_id)
    account = TradingAccount.objects.filter(
        user_id=user_id,
        pk=account_id,
        archived_at__isnull=True,
    ).first()
    if account is None:
        return {'ok': False, 'message': 'Compte de trading introuvable.'}

    if preferences.active_account_id == account.id:
        return {'ok': False, 'message': 'Le compte actif ne peut pas etre archive.'}

    available_accounts = list(get_trading_accounts_for_user(user_id))
    if len(available_accounts) <= 1:
        return {'ok': False, 'message': 'Au moins un compte actif doit etre conserve.'}

    account.archived_at = timezone.now()
    account.save(update_fields=['archived_at', 'updated_at'])

    return {
        'ok': True,
        'message': f'Compte archive : {build_account_label(account)}',
    }


def restore_trading_account_for_user(user_id, account_id):
    account = TradingAccount.objects.filter(
        user_id=user_id,
        pk=account_id,
        archived_at__isnull=False,
    ).first()
    if account is None:
        return {'ok': False, 'message': 'Compte archive introuvable.'}

    account.archived_at = None
    account.save(update_fields=['archived_at', 'updated_at'])
    return {
        'ok': True,
        'message': f'Compte restaure : {build_account_label(account)}',
    }


def filter_queryset_for_account(queryset, account):
    if account is None:
        return queryset
    return queryset.filter(models.Q(account=account) | models.Q(account__isnull=True))


def get_current_capital_for_user(user_id, preferences=None, account=None):
    preferences = preferences or get_or_create_preferences_for_user(user_id)
    account = account or get_or_create_active_account_for_user(user_id, preferences)
    trades = filter_queryset_for_account(
        Trade.objects.filter(user_id=user_id).order_by('executed_at', 'created_at'),
        account,
    )
    movements = filter_queryset_for_account(
        CapitalMovement.objects.filter(user_id=user_id).order_by('occurred_at', 'created_at'),
        account,
    )

    total_trade_pnl = sum((trade.net_pnl for trade in trades), Decimal('0.00'))
    total_deposits = sum(
        (movement.amount for movement in movements if movement.kind == CapitalMovement.Kind.DEPOSIT),
        Decimal('0.00'),
    )
    total_withdrawals = sum(
        (movement.amount for movement in movements if movement.kind == CapitalMovement.Kind.WITHDRAWAL),
        Decimal('0.00'),
    )
    return (account.capital_base + total_trade_pnl + total_deposits - total_withdrawals).quantize(Decimal('0.01'))


def serialize_preferences(preferences, current_capital=None, account=None, language=None):
    account = account or preferences.active_account
    capital_base = account.capital_base if account else preferences.capital_base
    currency = account.currency if account else preferences.currency
    currency_symbol = account.currency_symbol if account else preferences.currency_symbol
    current_capital = current_capital if current_capital is not None else capital_base
    return {
        'default_symbol': preferences.default_symbol,
        'default_direction': preferences.default_direction,
        'default_setup': preferences.default_setup,
        'default_lot_size': format_decimal_compact(preferences.default_lot_size),
        'default_gp_value': format_decimal_compact(preferences.default_fees),
        'default_fees': format_decimal_compact(preferences.default_fees),
        'default_confidence': preferences.default_confidence,
        'default_dashboard_year': str(preferences.default_dashboard_year),
        'default_week_start_day': str(normalize_week_start_day(getattr(preferences, 'default_week_start_day', calendar.SUNDAY))),
        'capital_base': format_decimal_compact(capital_base),
        'capital_base_formatted': format_currency(capital_base, currency),
        'current_capital': format_decimal_compact(current_capital),
        'current_capital_formatted': format_currency(current_capital, currency),
        'currency': currency,
        'currency_symbol': currency_symbol,
        'currency_label': TradingPreference.Currency(currency).label if currency else preferences.get_currency_display(),
        'active_account': {
            'id': account.id if account else None,
            'name': account.name if account else tr('common.main_account', language=language, default='Compte principal'),
            'broker': account.broker if account else '',
            'label': build_account_label(account, language=language),
        },
    }


def _local_datetime_for_excel(value):
    if not value:
        return None
    if timezone.is_aware(value):
        return timezone.localtime(value).replace(tzinfo=None)
    return value


def _normalize_export_cell_value(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return _local_datetime_for_excel(value)
    return value


def _autosize_export_columns(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value in (None, ''):
                continue
            if isinstance(cell.value, datetime):
                cell_length = len(cell.value.strftime('%Y-%m-%d %H:%M:%S'))
            else:
                cell_length = len(str(cell.value))
            max_length = max(max_length, cell_length)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def _append_export_sheet(workbook, title, headers, rows):
    from openpyxl.styles import Font, PatternFill

    worksheet = workbook.create_sheet(title=title)
    worksheet.append(headers)
    header_fill = PatternFill(fill_type='solid', fgColor='C8F800')
    header_font = Font(bold=True, color='06140F')

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        worksheet.append([_normalize_export_cell_value(value) for value in row])

    worksheet.freeze_panes = 'A2'
    _autosize_export_columns(worksheet)
    return worksheet


def _apply_export_scope_to_queryset(queryset, field_name, scope):
    start_datetime = scope.get('start_datetime')
    end_datetime = scope.get('end_datetime')
    if start_datetime is not None:
        queryset = queryset.filter(**{f'{field_name}__gte': start_datetime})
    if end_datetime is not None:
        queryset = queryset.filter(**{f'{field_name}__lte': end_datetime})
    return queryset


def _get_export_scope_datetimes(scope):
    start_date = scope.get('start_date')
    end_date = scope.get('end_date')
    if start_date is None or end_date is None:
        return None, None

    current_timezone = timezone.get_current_timezone()
    start_datetime = timezone.make_aware(datetime.combine(start_date, datetime.min.time()), current_timezone)
    end_datetime = timezone.make_aware(datetime.combine(end_date, datetime.max.time()), current_timezone)
    return start_datetime, end_datetime


def _safe_media_reference(file_field, base_url=None):
    if not file_field:
        return '', '', ''

    relative_path = (getattr(file_field, 'name', '') or '').replace('\\', '/')
    if not relative_path:
        return '', '', ''

    try:
        relative_url = file_field.url
    except ValueError:
        relative_url = ''

    absolute_url = urljoin(base_url, relative_url.lstrip('/')) if base_url and relative_url else relative_url
    return relative_path.split('/')[-1], relative_path, absolute_url


def build_user_data_export_workbook(user, scope, language=None, base_url=None):
    from openpyxl import Workbook

    scope = dict(scope or {})
    start_datetime, end_datetime = _get_export_scope_datetimes(scope)
    scope['start_datetime'] = start_datetime
    scope['end_datetime'] = end_datetime

    preferences = get_or_create_preferences_for_user(user.pk)
    active_account = get_or_create_active_account_for_user(user.pk, preferences)
    all_accounts = list(get_trading_accounts_for_user(user.pk, include_archived=True))
    trade_queryset = _apply_export_scope_to_queryset(
        Trade.objects.filter(user_id=user.pk).select_related('account').prefetch_related('screenshots').order_by('-executed_at', '-created_at'),
        'executed_at',
        scope,
    )
    movement_queryset = _apply_export_scope_to_queryset(
        CapitalMovement.objects.filter(user_id=user.pk).select_related('account').order_by('-occurred_at', '-created_at'),
        'occurred_at',
        scope,
    )
    trades = list(trade_queryset)
    movements = list(movement_queryset)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = 'Resume'

    summary_rows = [
        (tr('settings.export.summary.generated_at', language=language, default='Export genere le'), _local_datetime_for_excel(timezone.now())),
        (tr('settings.export.summary.username', language=language, default='Identifiant'), user.username),
        (tr('settings.export.summary.email', language=language, default='Email'), user.email or ''),
        (tr('settings.export.summary.active_account', language=language, default='Compte actif'), build_account_label(active_account, language=language)),
        (tr('settings.export.summary.scope', language=language, default='Filtre applique'), scope.get('label') or tr('settings.export.scope.all_time', language=language, default='Tout le temps')),
        (tr('settings.export.summary.period_start', language=language, default='Debut de periode'), scope.get('start_date') or tr('settings.export.summary.not_limited', language=language, default='Non limite')),
        (tr('settings.export.summary.period_end', language=language, default='Fin de periode'), scope.get('end_date') or tr('settings.export.summary.not_limited', language=language, default='Non limite')),
        (tr('settings.export.summary.accounts', language=language, default='Comptes exportes'), len(all_accounts)),
        (tr('settings.export.summary.trades', language=language, default='Trades exportes'), len(trades)),
        (tr('settings.export.summary.movements', language=language, default='Mouvements exportes'), len(movements)),
        (
            tr('settings.export.summary.screenshots', language=language, default='Captures exportees'),
            sum(len(serialize_trade_screenshots(trade)) for trade in trades),
        ),
        (
            tr('settings.export.summary.note', language=language, default='Note'),
            tr(
                'settings.export.summary.note_value',
                language=language,
                default='Le filtre de periode s applique aux trades, mouvements et captures. Les feuilles utilisateur, preferences et comptes restent completes.',
            ),
        ),
    ]
    summary_sheet.append(
        [
            tr('settings.export.summary.label_column', language=language, default='Champ'),
            tr('settings.export.summary.value_column', language=language, default='Valeur'),
        ]
    )
    for row in summary_rows:
        summary_sheet.append([_normalize_export_cell_value(value) for value in row])
    _append_export_sheet(
        workbook,
        'Utilisateur',
        [
            'id',
            'username',
            'first_name',
            'last_name',
            'email',
            'is_staff',
            'is_superuser',
            'date_joined',
            'last_login',
        ],
        [[
            user.pk,
            user.username,
            user.first_name or '',
            user.last_name or '',
            user.email or '',
            user.is_staff,
            user.is_superuser,
            _local_datetime_for_excel(user.date_joined),
            _local_datetime_for_excel(user.last_login),
        ]],
    )
    _append_export_sheet(
        workbook,
        'Preferences',
        [
            'user_id',
            'active_account_id',
            'active_account_name',
            'default_symbol',
            'default_direction',
            'default_setup',
            'default_lot_size',
            'default_risk_percent',
            'default_fees',
            'default_confidence',
            'capital_base',
            'currency',
            'ui_language',
            'default_dashboard_year',
            'default_week_start_day',
            'created_at',
            'updated_at',
        ],
        [[
            user.pk,
            preferences.active_account_id,
            build_account_label(preferences.active_account or active_account, language=language),
            preferences.default_symbol,
            preferences.default_direction,
            preferences.default_setup,
            preferences.default_lot_size,
            preferences.default_risk_percent,
            preferences.default_fees,
            preferences.default_confidence,
            preferences.capital_base,
            preferences.currency,
            preferences.ui_language,
            preferences.default_dashboard_year,
            preferences.default_week_start_day,
            _local_datetime_for_excel(preferences.created_at),
            _local_datetime_for_excel(preferences.updated_at),
        ]],
    )
    _append_export_sheet(
        workbook,
        'Comptes',
        [
            'id',
            'name',
            'broker',
            'account_identifier',
            'currency',
            'capital_base',
            'is_active',
            'is_archived',
            'created_at',
            'archived_at',
            'updated_at',
        ],
        [[
            account.pk,
            account.name,
            account.broker,
            account.account_identifier,
            account.currency,
            account.capital_base,
            account.pk == active_account.pk,
            account.is_archived,
            _local_datetime_for_excel(account.created_at),
            _local_datetime_for_excel(account.archived_at),
            _local_datetime_for_excel(account.updated_at),
        ] for account in all_accounts],
    )

    trade_rows = []
    screenshot_rows = []
    for trade in trades:
        legacy_name, legacy_path, legacy_url = _safe_media_reference(trade.screenshot, base_url=base_url)
        serialized_screenshots = serialize_trade_screenshots(trade)
        screenshot_names = []
        screenshot_urls = []

        if legacy_name:
            screenshot_names.append(legacy_name)
        if legacy_url:
            screenshot_urls.append(legacy_url)

        for screenshot in trade.screenshots.all():
            image_name, image_path, image_url = _safe_media_reference(screenshot.image, base_url=base_url)
            screenshot_names.append(image_name)
            if image_url:
                screenshot_urls.append(image_url)
            screenshot_rows.append(
                [
                    trade.pk,
                    trade.account_id,
                    trade.account.name if trade.account else '',
                    trade.symbol,
                    _local_datetime_for_excel(trade.executed_at),
                    'gallery',
                    screenshot.pk,
                    screenshot.sort_order,
                    image_name,
                    image_path,
                    image_url,
                    _local_datetime_for_excel(screenshot.created_at),
                ]
            )

        if legacy_name or legacy_path or legacy_url:
            screenshot_rows.append(
                [
                    trade.pk,
                    trade.account_id,
                    trade.account.name if trade.account else '',
                    trade.symbol,
                    _local_datetime_for_excel(trade.executed_at),
                    'legacy',
                    'legacy',
                    '',
                    legacy_name,
                    legacy_path,
                    legacy_url,
                    '',
                ]
            )

        trade_rows.append(
            [
                trade.pk,
                trade.user_id,
                trade.account_id,
                trade.account.name if trade.account else tr('common.main_account', language=language, default='Compte principal'),
                _local_datetime_for_excel(trade.executed_at),
                trade.symbol,
                trade.market,
                trade.direction,
                trade.result,
                trade.resolved_result,
                trade.setup,
                trade.entry_price,
                trade.rr_ratio,
                trade.exit_price,
                trade.quantity,
                trade.lot_size,
                trade.gp_value,
                trade.fees,
                trade.risk_amount,
                trade.risk_percent,
                trade.capital_base,
                trade.gross_pnl,
                trade.net_pnl,
                trade.risk_reward,
                trade.confidence,
                trade.notes,
                legacy_path,
                len(serialized_screenshots),
                ' | '.join(name for name in screenshot_names if name),
                ' | '.join(url for url in screenshot_urls if url),
                _local_datetime_for_excel(trade.created_at),
                _local_datetime_for_excel(trade.updated_at),
            ]
        )

    _append_export_sheet(
        workbook,
        'Trades',
        [
            'id',
            'user_id',
            'account_id',
            'account_name',
            'executed_at',
            'symbol',
            'market',
            'direction',
            'result',
            'resolved_result',
            'setup',
            'entry_price',
            'rr_ratio',
            'exit_price',
            'quantity',
            'lot_size',
            'gp_value',
            'fees',
            'risk_amount',
            'risk_percent',
            'capital_base',
            'gross_pnl',
            'net_pnl',
            'risk_reward',
            'confidence',
            'notes',
            'legacy_screenshot_path',
            'screenshot_count',
            'screenshot_names',
            'screenshot_urls',
            'created_at',
            'updated_at',
        ],
        trade_rows,
    )
    _append_export_sheet(
        workbook,
        'Mouvements',
        [
            'id',
            'user_id',
            'account_id',
            'account_name',
            'kind',
            'amount',
            'occurred_at',
            'note',
            'created_at',
            'updated_at',
        ],
        [[
            movement.pk,
            movement.user_id,
            movement.account_id,
            movement.account.name if movement.account else tr('common.main_account', language=language, default='Compte principal'),
            movement.kind,
            movement.amount,
            _local_datetime_for_excel(movement.occurred_at),
            movement.note,
            _local_datetime_for_excel(movement.created_at),
            _local_datetime_for_excel(movement.updated_at),
        ] for movement in movements],
    )
    _append_export_sheet(
        workbook,
        'Captures',
        [
            'trade_id',
            'account_id',
            'account_name',
            'symbol',
            'trade_executed_at',
            'source',
            'screenshot_id',
            'sort_order',
            'file_name',
            'file_path',
            'file_url',
            'created_at',
        ],
        screenshot_rows,
    )

    summary_sheet.freeze_panes = 'A2'
    _autosize_export_columns(summary_sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def attach_trade_screenshots(trade, uploaded_files):
    if not trade.pk:
        raise ValueError('Le trade doit etre sauvegarde avant l ajout des images.')

    files = [uploaded_file for uploaded_file in (uploaded_files or []) if uploaded_file]
    if not files:
        return

    current_max_sort_order = trade.screenshots.aggregate(max_sort_order=models.Max('sort_order'))['max_sort_order']
    start_index = 0 if current_max_sort_order is None else current_max_sort_order + 1
    for index, uploaded_file in enumerate(files, start=start_index):
        TradeScreenshot.objects.create(
            trade=trade,
            image=uploaded_file,
            sort_order=index,
        )


def serialize_trade_screenshots(trade):
    screenshots = []
    if trade.screenshot:
        screenshots.append(
            {
                'id': 'legacy',
                'url': trade.screenshot.url,
                'name': trade.screenshot.name.replace('\\', '/').split('/')[-1],
                'is_legacy': True,
            }
        )

    screenshots.extend(
        {
            'id': str(screenshot.pk),
            'url': screenshot.image.url,
            'name': screenshot.image.name.replace('\\', '/').split('/')[-1],
            'is_legacy': False,
        }
        for screenshot in trade.screenshots.all()
        if screenshot.image
    )
    return screenshots


def resequence_trade_screenshots(trade):
    screenshots = list(trade.screenshots.order_by('sort_order', 'pk'))
    for index, screenshot in enumerate(screenshots):
        if screenshot.sort_order != index:
            screenshot.sort_order = index

    if screenshots:
        TradeScreenshot.objects.bulk_update(screenshots, ['sort_order'])


def remove_trade_screenshots(trade, removed_ids):
    identifiers = [str(value).strip() for value in (removed_ids or []) if str(value).strip()]
    if not identifiers:
        return

    if 'legacy' in identifiers and trade.screenshot:
        trade.screenshot.delete(save=False)
        trade.screenshot = None
        trade.save(update_fields=['screenshot', 'updated_at'])

    screenshot_ids = []
    for identifier in identifiers:
        try:
            screenshot_ids.append(int(identifier))
        except (TypeError, ValueError):
            continue

    screenshots = list(trade.screenshots.filter(pk__in=screenshot_ids))
    for screenshot in screenshots:
        screenshot.image.delete(save=False)

    if screenshots:
        TradeScreenshot.objects.filter(pk__in=[screenshot.pk for screenshot in screenshots]).delete()
        resequence_trade_screenshots(trade)


def get_removed_screenshot_ids(payload):
    if hasattr(payload, 'getlist'):
        return payload.getlist('removed_screenshot_ids')
    if hasattr(payload, 'get'):
        value = payload.get('removed_screenshot_ids')
        if value is None:
            return []
        if isinstance(value, (list, tuple, set)):
            return list(value)
        return [value]
    return []


def serialize_trade(trade, currency='USD', language=None):
    executed_at = timezone.localtime(trade.executed_at)
    risk_reward = trade.risk_reward
    screenshots = serialize_trade_screenshots(trade)
    screenshot_urls = [item['url'] for item in screenshots]
    screenshot_url = screenshot_urls[0] if screenshot_urls else None
    serialized_result = trade.resolved_result
    serialized_gp_value = trade.gp_value
    if serialized_gp_value is None:
        serialized_gp_value = trade.net_pnl.quantize(Decimal('0.01'))
    serialized_ratio = trade.rr_ratio
    if serialized_ratio is None and trade.risk_amount not in (None, Decimal('0.00')):
        serialized_ratio = (trade.net_pnl / trade.risk_amount).quantize(Decimal('0.01'))
    result_label_map = {
        Trade.Result.TAKE_PROFIT: tr('trade.result.take_profit', language=language, default='Take profit'),
        Trade.Result.GAIN: tr('trade.result.gain', language=language, default='Gain'),
        Trade.Result.BREAK_EVEN: tr('trade.result.break_even', language=language, default='Break even'),
        Trade.Result.STOP_LOSS: tr('trade.result.stop_loss', language=language, default='Stoploss'),
        Trade.Result.LOSS: tr('trade.result.loss', language=language, default='Perte'),
    }
    capital_change_percent = (
        (trade.net_pnl / trade.capital_base * Decimal('100'))
        if trade.capital_base
        else None
    )
    return {
        'id': trade.id,
        'account_name': trade.account.name if trade.account else tr('common.main_account', language=language, default='Compte principal'),
        'symbol': trade.symbol,
        'direction': trade.get_direction_display(),
        'direction_code': trade.direction,
        'result': serialized_result,
        'result_code': serialized_result,
        'result_label': result_label_map.get(serialized_result, trade.resolved_result_label),
        'setup': trade.setup,
        'market': trade.market or 'Spot',
        'executed_at': executed_at.isoformat(),
        'executed_at_input': executed_at.strftime('%Y-%m-%dT%H:%M'),
        'executed_at_label': executed_at.strftime('%d/%m/%Y | %H:%M'),
        'entry_price': f'{trade.entry_price:,.4f}',
        'entry_price_value': f'{trade.entry_price:.4f}',
        'ratio': None if serialized_ratio is None else round(float(serialized_ratio), 2),
        'ratio_value': None if serialized_ratio is None else format_decimal_compact(abs(serialized_ratio)),
        'ratio_label': '--' if serialized_ratio is None else f'R {format_decimal_compact(serialized_ratio, use_grouping=True)}',
        'gp_value': None if serialized_gp_value is None else round(float(serialized_gp_value), 2),
        'gp_value_value': None if serialized_gp_value is None else format_decimal_compact(abs(serialized_gp_value)),
        'gp_value_label': '--' if serialized_gp_value is None else format_signed_value(serialized_gp_value),
        'lot_size': float(trade.lot_size or trade.quantity),
        'lot_size_value': format_decimal_compact(trade.lot_size or trade.quantity),
        'lot_size_label': f'{format_decimal_compact(trade.lot_size or trade.quantity, use_grouping=True)} lot(s)',
        'risk_percent': None if trade.risk_percent is None else round(float(trade.risk_percent), 2),
        'risk_percent_label': '--' if trade.risk_percent is None else f'{format_decimal_compact(trade.risk_percent, use_grouping=True)}%',
        'capital_change_percent': None if capital_change_percent is None else round(float(capital_change_percent), 2),
        'capital_change_percent_label': '--' if capital_change_percent is None else format_signed_percent(capital_change_percent),
        'risk_amount_formatted': '--' if trade.risk_amount is None else format_currency(trade.risk_amount, currency),
        'capital_base_value': format_decimal_compact(trade.capital_base),
        'capital_base_formatted': format_currency(trade.capital_base, currency),
        'fees_formatted': format_currency(trade.fees, currency),
        'pnl': float(trade.net_pnl),
        'pnl_formatted': format_currency(trade.net_pnl, currency),
        'pnl_tone': 'profit' if trade.net_pnl >= 0 else 'loss',
        'risk_reward': None if risk_reward is None else round(float(risk_reward), 2),
        'confidence': trade.confidence,
        'confidence_label': f'{trade.confidence}/5',
        'notes': trade.notes,
        'screenshots': screenshots,
        'screenshot_url': screenshot_url,
        'screenshot_urls': screenshot_urls,
        'screenshot_count': len(screenshot_urls),
    }


def build_calendar_payload(year, month, daily_totals, daily_counts, trades_by_day, currency='USD', language=None, firstweekday=calendar.SUNDAY):
    firstweekday = normalize_week_start_day(firstweekday)
    today = timezone.localdate()
    month_calendar = calendar.Calendar(firstweekday=firstweekday).monthdatescalendar(year, month)
    rows = []
    week_summaries = []

    for index, week in enumerate(month_calendar, start=1):
        row = []
        week_pnl = 0.0
        active_days = 0
        for day in week:
            pnl = round(daily_totals.get(day, 0.0), 2)
            trade_count = daily_counts.get(day, 0)
            if day.month == month and trade_count:
                week_pnl += pnl
                active_days += 1
            row.append(
                {
                    'iso': day.isoformat(),
                    'day': day.day,
                    'is_today': day == today,
                    'in_month': day.month == month,
                    'pnl': pnl,
                    'pnl_formatted': format_currency(pnl, currency) if pnl else format_currency(0, currency),
                    'trade_count': trade_count,
                    'tone': 'profit' if pnl > 0 else 'loss' if pnl < 0 else 'flat',
                }
            )
        rows.append(row)
        week_summaries.append(
            {
                'label': tr('calendar.week', language=language, default='Semaine {index}', index=index),
                'pnl': week_pnl,
                'pnl_formatted': format_currency(week_pnl, currency),
                'active_days': active_days,
                'tone': 'profit' if week_pnl > 0 else 'loss' if week_pnl < 0 else 'flat',
            }
        )

    return {
        'label': format_month_label(year, month, language),
        'weekday_labels': get_rotated_weekday_short_labels(language, firstweekday=firstweekday),
        'week_start_day': firstweekday,
        'rows': rows,
        'week_summaries': week_summaries,
        'trade_map': {
            day.isoformat(): [serialize_trade(trade, currency, language=language) for trade in day_trades]
            for day, day_trades in trades_by_day.items()
        },
    }


def get_demo_dataset_state_for_user(user_id, account=None):
    preferences = None
    active_account = account
    if active_account is None:
        preferences = get_or_create_preferences_for_user(user_id)
        active_account = get_or_create_active_account_for_user(user_id, preferences)

    account_trades = filter_queryset_for_account(
        Trade.objects.filter(user_id=user_id),
        active_account,
    )
    demo_trade_count = account_trades.filter(notes=DEMO_TRADE_NOTE).count()
    action = 'unload' if demo_trade_count else 'load' if not account_trades.exists() else 'hidden'
    return {
        'loaded': bool(demo_trade_count),
        'trade_count': demo_trade_count,
        'action': action,
    }


def build_dashboard_payload_for_user(user_id, raw_month=None, raw_year=None, language=None):
    language = normalize_language(language)
    preferences = get_or_create_preferences_for_user(user_id)
    active_account = get_or_create_active_account_for_user(user_id, preferences)
    week_start_day = normalize_week_start_day(getattr(preferences, 'default_week_start_day', calendar.SUNDAY))
    currency = active_account.currency
    trades = list(
        filter_queryset_for_account(
            Trade.objects.filter(user_id=user_id),
            active_account,
        )
        .select_related('user')
        .prefetch_related('screenshots')
        .order_by('executed_at', 'created_at')
    )
    movements = list(
        filter_queryset_for_account(
            CapitalMovement.objects.filter(user_id=user_id),
            active_account,
        )
        .select_related('user')
        .order_by('occurred_at', 'created_at')
    )
    available_years = get_available_dashboard_years(trades)
    year = resolve_selected_year(raw_year, raw_month, preferences, available_years)
    year_trades = [
        trade for trade in trades
        if timezone.localtime(trade.executed_at).year == year
    ]
    month = resolve_selected_month(raw_month, year, year_trades)

    filtered_trades = []
    for trade in year_trades:
        local_executed_at = timezone.localtime(trade.executed_at)
        if local_executed_at.month == month:
            filtered_trades.append(trade)

    available_months = []
    seen_months = set()
    for trade in sorted(year_trades, key=lambda value: value.executed_at, reverse=True):
        local_date = timezone.localtime(trade.executed_at)
        key = f'{year}-{local_date.month:02d}'
        if key in seen_months:
            continue
        seen_months.add(key)
        available_months.append(
            {
                'value': key,
                'label': format_month_label(year, local_date.month, language),
            }
        )

    current_month_value = f'{year}-{month:02d}'
    if current_month_value not in seen_months:
        available_months.insert(
            0,
            {
                'value': current_month_value,
                'label': format_month_label(year, month, language),
            },
        )
    if not available_months:
        available_months.append(
            {
                'value': current_month_value,
                'label': format_month_label(year, month, language),
            }
        )

    available_year_options = [
        {
            'value': str(available_year),
            'label': str(available_year),
        }
        for available_year in available_years
    ]

    monthly_pnls = [float(trade.net_pnl) for trade in filtered_trades]
    winners = [value for value in monthly_pnls if value > 0]
    losers = [value for value in monthly_pnls if value < 0]
    total_net = sum(monthly_pnls)
    trade_count = len(filtered_trades)
    win_rate = (len(winners) / trade_count * 100) if trade_count else 0
    positive_sum = sum(winners)
    negative_sum = abs(sum(losers))
    profit_factor = positive_sum / negative_sum if negative_sum else (positive_sum if positive_sum else 0)
    avg_trade = total_net / trade_count if trade_count else 0

    daily_totals = defaultdict(float)
    daily_counts = defaultdict(int)
    setup_performance = defaultdict(float)
    trades_by_day = defaultdict(list)
    for trade in filtered_trades:
        trade_day = timezone.localtime(trade.executed_at).date()
        pnl_value = float(trade.net_pnl)
        daily_totals[trade_day] += pnl_value
        daily_counts[trade_day] += 1
        setup_performance[trade.setup] += pnl_value
        trades_by_day[trade_day].append(trade)

    active_days = len(daily_counts)
    profitable_days = len([value for value in daily_totals.values() if value > 0])
    profitable_day_rate = (profitable_days / active_days * 100) if active_days else 0

    sorted_days = sorted(daily_totals.items(), key=lambda item: item[0])
    cumulative_values = []
    cumulative_total = 0.0
    for _, value in sorted_days:
        cumulative_total += value
        cumulative_values.append(round(cumulative_total, 2))

    max_drawdown = compute_drawdown(cumulative_values)
    recovery_factor = total_net / abs(max_drawdown) if max_drawdown else total_net if total_net > 0 else 0
    avg_win = (sum(winners) / len(winners)) if winners else 0
    avg_loss = abs(sum(losers) / len(losers)) if losers else 0
    avg_win_loss_ratio = avg_win / avg_loss if avg_loss else (avg_win if avg_win else 0)

    daily_values = list(daily_totals.values())
    if len(daily_values) > 1:
        average_abs = sum(abs(value) for value in daily_values) / len(daily_values) or 1
        consistency_score = clamp(100 - ((pstdev(daily_values) / average_abs) * 30))
    else:
        consistency_score = 60 if daily_values else 0

    profit_factor_score = 100 if profit_factor >= 3 else clamp((profit_factor / 3) * 100)
    avg_win_loss_score = 100 if avg_win_loss_ratio >= 3 else clamp((avg_win_loss_ratio / 3) * 100)
    drawdown_score = 100 if max_drawdown == 0 else clamp(100 - abs(max_drawdown) / max(abs(total_net), 1) * 60)
    recovery_score = 100 if recovery_factor >= 4 else clamp((max(recovery_factor, 0) / 4) * 100)
    if not filtered_trades:
        overall_score = 0
        drawdown_score = 0
    else:
        overall_score = round(
            (
                win_rate
                + consistency_score
                + profit_factor_score
                + avg_win_loss_score
                + drawdown_score
                + recovery_score
            )
            / 6
        )

    if filtered_trades:
        best_setup_name = max(setup_performance.items(), key=lambda item: item[1])[0]
        best_day_value = max(daily_totals.values()) if daily_totals else 0
        worst_day_value = min(daily_totals.values()) if daily_totals else 0
        insight_lines = [
            tr('dashboard.insight.best_setup', language=language, default='Setup le plus performant : {setup}', setup=best_setup_name),
            tr('dashboard.insight.best_day', language=language, default='Meilleure journee : {value}', value=format_currency(best_day_value, currency)),
            tr('dashboard.insight.max_drawdown', language=language, default='Drawdown quotidien maximal : {value}', value=format_currency(worst_day_value, currency)),
        ]
    else:
        insight_lines = [
            tr('dashboard.insight.empty_one', language=language, default='Enregistrez les premiers trades pour activer les indicateurs.'),
        ]

    monthly_trades = [serialize_trade(trade, currency, language=language) for trade in reversed(filtered_trades)]
    recent_trades = monthly_trades[:5]

    metric_cards = [
        {
            'key': 'net_pnl',
            'label': tr('dashboard.metric.net_pnl', language=language, default='Net P&L'),
            'value': format_currency(total_net, currency),
            'detail': tr('dashboard.metric.net_pnl_detail', language=language, default='{count} trade(s) sur la periode selectionnee', count=trade_count),
            'progress': clamp(50 + total_net / 100),
            'tone': 'profit' if total_net >= 0 else 'loss',
        },
        {
            'key': 'win_rate',
            'label': tr('dashboard.metric.win_rate', language=language, default='Taux de reussite'),
            'value': f'{win_rate:.2f}%',
            'detail': tr('dashboard.metric.win_rate_detail', language=language, default='{winners} trade(s) gagnants / {count}', winners=len(winners), count=trade_count),
            'progress': clamp(win_rate),
            'tone': 'profit' if win_rate >= 50 else 'loss',
        },
        {
            'key': 'profit_factor',
            'label': tr('dashboard.metric.profit_factor', language=language, default='Profit factor'),
            'value': f'{profit_factor:.2f}',
            'detail': tr('dashboard.metric.profit_factor_detail', language=language, default='Ratio gains / pertes'),
            'progress': profit_factor_score,
            'tone': 'profit' if profit_factor >= 1.5 else 'loss',
        },
        {
            'key': 'day_rate',
            'label': tr('dashboard.metric.positive_days', language=language, default='Jours positifs'),
            'value': f'{profitable_day_rate:.2f}%',
            'detail': tr('dashboard.metric.positive_days_detail', language=language, default='{days} jour(s) positifs / {active_days}', days=profitable_days, active_days=active_days),
            'progress': clamp(profitable_day_rate),
            'tone': 'profit' if profitable_day_rate >= 50 else 'loss',
        },
        {
            'key': 'avg_trade',
            'label': tr('dashboard.metric.avg_trade', language=language, default='Moyenne / trade'),
            'value': format_currency(avg_trade, currency),
            'detail': tr('dashboard.metric.avg_trade_detail', language=language, default='Ratio moyen gain / perte : {ratio}', ratio=f'{avg_win_loss_ratio:.2f}'),
            'progress': clamp(50 + avg_trade / 20),
            'tone': 'profit' if avg_trade >= 0 else 'loss',
        },
    ]

    all_time_net_decimal = sum((trade.net_pnl for trade in trades), Decimal('0.00'))
    all_time_net = float(all_time_net_decimal)
    total_deposits = sum(
        (movement.amount for movement in movements if movement.kind == CapitalMovement.Kind.DEPOSIT),
        Decimal('0.00'),
    )
    total_withdrawals = sum(
        (movement.amount for movement in movements if movement.kind == CapitalMovement.Kind.WITHDRAWAL),
        Decimal('0.00'),
    )
    current_capital = active_account.capital_base + all_time_net_decimal + total_deposits - total_withdrawals
    demo_state = get_demo_dataset_state_for_user(user_id, account=active_account)

    return {
        'summary': {
            'selected_year': str(year),
            'selected_year_label': str(year),
            'selected_month': current_month_value,
            'selected_month_label': format_month_label(year, month, language),
            'has_data': bool(filtered_trades),
            'trade_count': trade_count,
        },
        'overview': {
            'all_time_pnl': format_currency(all_time_net, currency),
            'all_time_trade_count': len(trades),
            'active_days': active_days,
            'best_setup': max(setup_performance, key=setup_performance.get) if setup_performance else '--',
            'score': overall_score,
        },
        'available_years': available_year_options,
        'available_months': available_months,
        'demo': demo_state,
        'metrics': metric_cards,
        'scorecard': {
            'value': overall_score,
            'caption': tr('dashboard.score.caption', language=language, default='Indice de discipline et de performance'),
            'insights': insight_lines,
        },
        'charts': {
            'radar': {
                'labels': [
                    tr('dashboard.chart.win_rate', language=language, default='Taux de reussite'),
                    tr('dashboard.chart.consistency', language=language, default='Regularite'),
                    tr('dashboard.metric.profit_factor', language=language, default='Profit factor'),
                    tr('dashboard.chart.gain_loss', language=language, default='Gain / perte'),
                    tr('dashboard.chart.recovery', language=language, default='Recuperation'),
                    tr('dashboard.chart.drawdown', language=language, default='Drawdown'),
                ],
                'values': [
                    round(win_rate, 2),
                    round(consistency_score, 2),
                    round(profit_factor_score, 2),
                    round(avg_win_loss_score, 2),
                    round(recovery_score, 2),
                    round(drawdown_score, 2),
                ],
            },
            'cumulative': {
                'labels': [format_short_day(day) for day, _ in sorted_days],
                'values': cumulative_values,
            },
            'daily': {
                'labels': [format_short_day(day) for day, _ in sorted_days],
                'values': [round(value, 2) for _, value in sorted_days],
            },
            'combo': {
                'labels': [format_short_day(day) for day, _ in sorted_days],
                'daily': [round(value, 2) for _, value in sorted_days],
                'cumulative': cumulative_values,
            },
        },
        'calendar': build_calendar_payload(
            year,
            month,
            daily_totals,
            daily_counts,
            trades_by_day,
            currency,
            language=language,
            firstweekday=week_start_day,
        ),
        'recent_trades': recent_trades,
        'monthly_trades': monthly_trades,
        'preferences': serialize_preferences(preferences, current_capital=current_capital, account=active_account, language=language),
    }


def create_trade_for_user(user_id, payload, files, form_class, language=None):
    user = get_user_model().objects.get(pk=user_id)
    preferences = get_or_create_preferences_for_user(user_id)
    active_account = get_or_create_active_account_for_user(user_id, preferences)
    current_capital = get_current_capital_for_user(user_id, preferences, account=active_account)
    form = form_class(
        payload,
        files,
        preferences=preferences,
        capital_base_override=current_capital,
        language=language,
    )
    if not form.is_valid():
        return {'ok': False, 'errors': form.errors.get_json_data()}

    with transaction.atomic():
        trade = form.save(commit=False)
        trade.user = user
        trade.account = active_account
        trade.save()
        attach_trade_screenshots(trade, form.cleaned_data.get('screenshots'))
    return {'ok': True, 'trade': serialize_trade(trade, active_account.currency, language=language)}


def update_trade_for_user(user_id, trade_id, payload, files, form_class, language=None):
    preferences = get_or_create_preferences_for_user(user_id)
    active_account = get_or_create_active_account_for_user(user_id, preferences)
    trade = filter_queryset_for_account(
        Trade.objects.filter(user_id=user_id, pk=trade_id),
        active_account,
    ).first()
    if not trade:
        return {'ok': False, 'message': tr('dashboard.modal.trade', language=language, default='Trade') + ' introuvable.'}

    form = form_class(payload, files, instance=trade, preferences=preferences, language=language)
    if not form.is_valid():
        return {'ok': False, 'errors': form.errors.get_json_data()}

    with transaction.atomic():
        updated_trade = form.save(commit=False)
        updated_trade.user_id = user_id
        if updated_trade.account_id is None:
            updated_trade.account = active_account
        updated_trade.save()
        remove_trade_screenshots(updated_trade, get_removed_screenshot_ids(payload))
        attach_trade_screenshots(updated_trade, form.cleaned_data.get('screenshots'))
    trade_currency = updated_trade.account.currency if updated_trade.account else active_account.currency
    return {'ok': True, 'trade': serialize_trade(updated_trade, trade_currency, language=language)}


def serialize_capital_movement(movement, currency='USD', language=None):
    occurred_at = timezone.localtime(movement.occurred_at)
    signed_amount = movement.amount if movement.kind == CapitalMovement.Kind.DEPOSIT else -movement.amount
    kind_label = (
        tr('transactions.kind.deposit', language=language, default='Depot')
        if movement.kind == CapitalMovement.Kind.DEPOSIT
        else tr('transactions.kind.withdrawal', language=language, default='Retrait')
    )
    return {
        'id': movement.id,
        'account_name': movement.account.name if movement.account else tr('common.main_account', language=language, default='Compte principal'),
        'kind': movement.kind,
        'kind_label': kind_label,
        'amount': float(movement.amount),
        'amount_label': format_currency(signed_amount, currency),
        'occurred_at': occurred_at.isoformat(),
        'occurred_at_label': occurred_at.strftime('%d/%m/%Y | %H:%M'),
        'note': movement.note,
        'tone': 'deposit' if movement.kind == CapitalMovement.Kind.DEPOSIT else 'withdrawal',
    }


def build_transactions_payload_for_user(user_id, language=None):
    language = normalize_language(language)
    preferences = get_or_create_preferences_for_user(user_id)
    active_account = get_or_create_active_account_for_user(user_id, preferences)
    currency = active_account.currency
    trades = list(
        filter_queryset_for_account(
            Trade.objects.filter(user_id=user_id),
            active_account,
        )
        .select_related('user')
        .prefetch_related('screenshots')
        .order_by('executed_at', 'created_at')
    )
    movements = list(
        filter_queryset_for_account(
            CapitalMovement.objects.filter(user_id=user_id),
            active_account,
        )
        .select_related('user')
        .order_by('occurred_at', 'created_at')
    )

    monthly = {}

    def ensure_month(month_start):
        if month_start not in monthly:
            monthly[month_start] = {
                'month_start': month_start,
                'gp_total': Decimal('0.00'),
                'deposits': Decimal('0.00'),
                'withdrawals': Decimal('0.00'),
                'trade_count': 0,
                'winners': 0,
                'losers': 0,
            }
        return monthly[month_start]

    current_month_start = get_month_start(timezone.localdate())
    ensure_month(current_month_start)

    total_trade_pnl = Decimal('0.00')
    for trade in trades:
        local_dt = timezone.localtime(trade.executed_at)
        month_bucket = ensure_month(get_month_start(local_dt.date()))
        pnl = trade.net_pnl
        total_trade_pnl += pnl
        month_bucket['gp_total'] += pnl
        month_bucket['trade_count'] += 1
        if pnl > 0:
            month_bucket['winners'] += 1
        elif pnl < 0:
            month_bucket['losers'] += 1

    total_deposits = Decimal('0.00')
    total_withdrawals = Decimal('0.00')
    for movement in movements:
        local_dt = timezone.localtime(movement.occurred_at)
        month_bucket = ensure_month(get_month_start(local_dt.date()))
        if movement.kind == CapitalMovement.Kind.DEPOSIT:
            month_bucket['deposits'] += movement.amount
            total_deposits += movement.amount
        else:
            month_bucket['withdrawals'] += movement.amount
            total_withdrawals += movement.amount

    sorted_months_desc = sorted(monthly.values(), key=lambda item: item['month_start'], reverse=True)
    current_month = monthly[current_month_start]
    current_capital = active_account.capital_base + total_trade_pnl + total_deposits - total_withdrawals
    current_month_net = current_month['gp_total'] + current_month['deposits'] - current_month['withdrawals']

    comparable_months = [
        item for item in sorted_months_desc
        if item['trade_count'] or item['deposits'] or item['withdrawals']
    ]

    def resolve_month_net(month_data):
        return month_data['gp_total'] + month_data['deposits'] - month_data['withdrawals']

    def resolve_month_progress_percent(month_data):
        capital_start = month_data.get('capital_start')
        if not capital_start:
            return None

        month_net = resolve_month_net(month_data)
        return (month_net / capital_start * Decimal('100')).quantize(Decimal('0.01'))

    best_month = max(comparable_months, key=resolve_month_net) if comparable_months else None
    completed_comparable_months = [
        item for item in comparable_months
        if item['month_start'] != current_month_start
    ]
    completed_loss_months = [
        item for item in completed_comparable_months
        if resolve_month_net(item) < 0
    ]
    worst_month = min(completed_loss_months, key=resolve_month_net) if completed_loss_months else None

    running_capital = active_account.capital_base
    for month_data in sorted(monthly.values(), key=lambda item: item['month_start']):
        month_data['capital_start'] = running_capital
        running_capital += resolve_month_net(month_data)

    monthly_history = []
    for month_data in sorted_months_desc:
        month_net = month_data['gp_total'] + month_data['deposits'] - month_data['withdrawals']
        progress_percent = resolve_month_progress_percent(month_data)
        monthly_history.append(
            {
                'month_key': month_data['month_start'].strftime('%Y-%m'),
                'month_label': format_month_label(month_data['month_start'].year, month_data['month_start'].month, language),
                'capital_start': float(month_data['capital_start']),
                'capital_start_label': format_currency(month_data['capital_start'], currency),
                'gp_total': float(month_data['gp_total']),
                'gp_total_label': format_currency(month_data['gp_total'], currency),
                'deposits_label': format_currency(month_data['deposits'], currency),
                'withdrawals_label': format_currency(month_data['withdrawals'], currency),
                'trade_count': month_data['trade_count'],
                'winners': month_data['winners'],
                'losers': month_data['losers'],
                'net_label': format_currency(month_net, currency),
                'progress_label': '--' if progress_percent is None else format_signed_percent(progress_percent),
                'progress_tone': 'profit' if progress_percent and progress_percent > 0 else 'loss' if progress_percent and progress_percent < 0 else 'flat',
                'tone': 'profit' if month_net > 0 else 'loss' if month_net < 0 else 'flat',
                'is_best_month': bool(best_month and month_data['month_start'] == best_month['month_start']),
                'is_worst_month': bool(worst_month and month_data['month_start'] == worst_month['month_start']),
            }
        )

    chart_months = list(reversed(sorted_months_desc[:12]))
    all_movements = [serialize_capital_movement(movement, currency, language=language) for movement in reversed(movements)]
    recent_movements = all_movements[:8]

    return {
        'summary': {
            'current_month_label': format_month_label(current_month_start.year, current_month_start.month, language),
            'current_capital_label': format_currency(current_capital, currency),
            'base_capital_label': format_currency(active_account.capital_base, currency),
            'trade_count_month': current_month['trade_count'],
            'winners_month': current_month['winners'],
            'losers_month': current_month['losers'],
        },
        'metrics': [
            {
                'label': tr('transactions.metric.current_capital', language=language, default='Capital actuel'),
                'value': format_currency(current_capital, currency),
                'detail': tr('transactions.metric.current_capital_detail', language=language, default='Capital initial {capital} + performances + flux', capital=format_currency(active_account.capital_base, currency)),
                'progress': clamp(55 + float(total_trade_pnl) / 100),
                'tone': 'profit' if current_capital >= active_account.capital_base else 'loss',
            },
            {
                'label': tr('transactions.metric.trades_month', language=language, default='Trades du mois'),
                'value': str(current_month['trade_count']),
                'detail': tr('transactions.metric.trades_month_detail', language=language, default='{winners} gagnant(s) / {losers} perdant(s)', winners=current_month['winners'], losers=current_month['losers']),
                'progress': clamp(current_month['trade_count'] * 8, 0, 100),
                'tone': 'profit' if current_month['winners'] >= current_month['losers'] else 'loss',
            },
            {
                'label': tr('transactions.metric.gp_month', language=language, default='G/P du mois'),
                'value': format_currency(current_month['gp_total'], currency),
                'detail': format_month_label(current_month_start.year, current_month_start.month, language),
                'progress': clamp(50 + float(current_month['gp_total']) / 100),
                'tone': 'profit' if current_month['gp_total'] >= 0 else 'loss',
            },
            {
                'label': tr('transactions.metric.withdrawals', language=language, default='Retraits'),
                'value': format_currency(total_withdrawals, currency),
                'detail': tr('transactions.metric.withdrawals_detail', language=language, default='{count} operation(s) enregistree(s)', count=len([item for item in movements if item.kind == CapitalMovement.Kind.WITHDRAWAL])),
                'progress': clamp(float(total_withdrawals) / 50),
                'tone': 'profit',
            },
            {
                'label': tr('transactions.metric.deposits', language=language, default='Depots'),
                'value': format_currency(total_deposits, currency),
                'detail': tr('transactions.metric.deposits_detail', language=language, default='{count} operation(s) enregistree(s)', count=len([item for item in movements if item.kind == CapitalMovement.Kind.DEPOSIT])),
                'progress': clamp(float(total_deposits) / 50),
                'tone': 'amber',
            },
        ],
        'chart': {
            'labels': [format_month_label(item['month_start'].year, item['month_start'].month, language) for item in chart_months],
            'values': [round(float(item['gp_total']), 2) for item in chart_months],
        },
        'monthly_history': monthly_history,
        'all_movements': all_movements,
        'recent_movements': recent_movements,
        'highlights': {
            'net_month_label': format_currency(current_month_net, currency),
            'deposits_label': format_currency(current_month['deposits'], currency),
            'withdrawals_label': format_currency(current_month['withdrawals'], currency),
            'best_month_label': (
                f"{format_month_label(best_month['month_start'].year, best_month['month_start'].month, language)} | "
                f"{format_currency(resolve_month_net(best_month), currency)}"
            ) if best_month else '--',
            'worst_month_label': (
                f"{format_month_label(worst_month['month_start'].year, worst_month['month_start'].month, language)} | "
                f"{format_currency(resolve_month_net(worst_month), currency)}"
            ) if worst_month else '--',
        },
    }


def create_capital_movement_for_user(user_id, payload, form_class, language=None):
    preferences = get_or_create_preferences_for_user(user_id)
    active_account = get_or_create_active_account_for_user(user_id, preferences)
    user = get_user_model().objects.get(pk=user_id)
    form = form_class(payload)
    if not form.is_valid():
        return {'ok': False, 'errors': form.errors.get_json_data()}

    movement = form.save(commit=False)
    movement.user = user
    movement.account = active_account
    movement.save()
    kind_label = (
        tr('transactions.kind.deposit', language=language, default='Depot')
        if movement.kind == CapitalMovement.Kind.DEPOSIT
        else tr('transactions.kind.withdrawal', language=language, default='Retrait')
    )
    return {
        'ok': True,
        'message': tr('transactions.movement.saved', language=language, default='{kind} enregistre.', kind=kind_label),
        'movement': serialize_capital_movement(movement, active_account.currency, language=language),
    }


def seed_demo_trades_for_user(user_id, language=None):
    preferences = get_or_create_preferences_for_user(user_id)
    active_account = get_or_create_active_account_for_user(user_id, preferences)
    if filter_queryset_for_account(Trade.objects.filter(user_id=user_id), active_account).exists():
        return {'ok': False, 'message': tr('transactions.demo.exists', language=language, default='Des trades sont deja enregistres sur ce compte.')}

    user = get_user_model().objects.get(pk=user_id)
    now = timezone.localtime()
    base_month = date(now.year, now.month, 1)
    randomizer = random.Random(42)

    demo_trades = []
    setups = ['Breakout', 'Reversal', 'VWAP reclaim', 'London impulse']
    symbols = ['EURUSD', 'XAUUSD', 'NAS100', 'BTCUSD']
    markets = {
        'EURUSD': 'Forex',
        'XAUUSD': 'Commodities',
        'NAS100': 'Indices',
        'BTCUSD': 'Crypto',
    }

    for offset in range(18):
        day = base_month + timedelta(days=min(offset * 2, 25))
        entry = Decimal(str(randomizer.uniform(10, 200))).quantize(Decimal('0.0001'))
        move = Decimal(str(randomizer.uniform(-7, 14))).quantize(Decimal('0.0001'))
        quantity = Decimal(str(randomizer.uniform(2, 9))).quantize(Decimal('0.01'))
        risk_percent = Decimal(str(randomizer.uniform(0.30, 1.80))).quantize(Decimal('0.01'))
        risk_amount = (active_account.capital_base * risk_percent / Decimal('100')).quantize(Decimal('0.01'))
        rr_ratio = Decimal(str(randomizer.uniform(-1.40, 3.60))).quantize(Decimal('0.01'))
        result = (
            Trade.Result.TAKE_PROFIT if rr_ratio > 0 else
            Trade.Result.STOP_LOSS if rr_ratio < 0 else
            Trade.Result.BREAK_EVEN
        )
        gp_value = (abs(rr_ratio) * Decimal(str(randomizer.uniform(10, 80)))).quantize(Decimal('0.01'))
        if result == Trade.Result.STOP_LOSS:
            gp_value = -gp_value
        elif result == Trade.Result.BREAK_EVEN:
            gp_value = Decimal('0.00')
        symbol = symbols[offset % len(symbols)]
        executed_time = datetime.min.time().replace(hour=9 + (offset % 6), minute=15)
        executed_at = timezone.make_aware(datetime.combine(day, executed_time))
        trade = Trade(
            user=user,
            account=active_account,
            executed_at=executed_at,
            symbol=symbol,
            market=markets[symbol],
            direction=Trade.Direction.LONG if offset % 3 else Trade.Direction.SHORT,
            result=result,
            setup=setups[offset % len(setups)],
            entry_price=entry,
            rr_ratio=rr_ratio,
            exit_price=(entry + move).quantize(Decimal('0.0001')),
            quantity=quantity,
            lot_size=quantity,
            gp_value=gp_value,
            fees=Decimal('4.50'),
            risk_amount=risk_amount,
            risk_percent=risk_percent,
            capital_base=active_account.capital_base,
            confidence=(offset % 5) + 1,
            notes=DEMO_TRADE_NOTE,
        )
        demo_trades.append(trade)

    with transaction.atomic():
        Trade.objects.bulk_create(demo_trades)

    return {'ok': True, 'message': tr('transactions.demo.loaded', language=language, default='Jeu de donnees de demonstration charge.')}


def clear_demo_trades_for_user(user_id, language=None):
    preferences = get_or_create_preferences_for_user(user_id)
    active_account = get_or_create_active_account_for_user(user_id, preferences)
    demo_trades = filter_queryset_for_account(
        Trade.objects.filter(user_id=user_id, notes=DEMO_TRADE_NOTE),
        active_account,
    )

    if not demo_trades.exists():
        return {
            'ok': False,
            'message': tr(
                'transactions.demo.missing',
                language=language,
                default='Aucune donnee de demonstration a decharger sur ce compte.',
            ),
        }

    with transaction.atomic():
        demo_trades.delete()

    return {
        'ok': True,
        'message': tr(
            'transactions.demo.unloaded',
            language=language,
            default='Jeu de donnees de demonstration decharge.',
        ),
    }
